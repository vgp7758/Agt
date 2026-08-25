"""real_tools.py —— 真实强力工具（Step 6）。

把玩具计算器升级为能干实事的工具集：
  run_python              : 执行模型写的 Python（独立子进程 + 超时）。
  read_file/write_file/list_dir : 读写文件，限定在 workspace/ 内（控制爆炸半径）。
  web_search              : 联网搜索（DuckDuckGo，无需 key；国内可能需代理）。
  open_url                : 抓取网页提取正文文本（start/max_chars 分页续读）。
  run_shell               : 执行系统命令（最强大也最危险，超时 + 日志）。

安全策略：
  - 代码/命令在独立子进程中执行，带超时，超时即终止，不会卡死 Agent。
  - 文件操作限定在 workspace/ 目录，越界拒绝，防误伤系统文件。
  - 任何工具出错都转成文本回传模型（不抛异常炸流程）。
"""
from __future__ import annotations

import base64
import hashlib
import mimetypes
import os
import queue
import re
import subprocess
import sys
import tempfile
import threading
import time
from pathlib import Path

from tools import Tool, Toolbox

# 工作区 = 启动时的当前目录(cwd)。文件读写、代码执行都以这里为根与沙箱边界。
# 故可从任意目录 `python /path/to/chat.py` 启动，在当前目录执行任务。
WORKSPACE = Path.cwd()

# 代码执行与 shell 的超时秒数（可通过 set_timeout 工具运行时调整）
TOOL_TIMEOUT = 10

# 工具执行进度回调（由 agent 在执行工具前设置；流式输出/心跳通过它推给 UI）
_tool_emit = None

# 后台任务表：超时未完成的 run_python/run_shell 子进程转后台后注册在此
_bg_tasks: dict = {}


def _resolve(path: str) -> Path:
    """把路径解析到 workspace 内；越界则抛 PermissionError（会被 Tool.run 转成文本）。"""
    base = WORKSPACE.resolve()
    target = (base / path).resolve()
    try:
        target.relative_to(base)  # 不在 base 下会抛 ValueError
    except ValueError:
        raise PermissionError(f"拒绝访问 workspace 外的路径: {path}")
    return target


# ===== 文件版本（乐观锁）=====
# 行号类编辑工具（insert/delete/move）没有 old_string 自校验，靠 version 防止
# "模型脑子里的行号已过期 → 静默写错位置"。模型 read_file/grep 时拿到当前 version，
# 编辑时原样回传；服务端重新算一次当前 version 比对：相等→安全应用，不等→拒绝要求重读。
# 服务端无需持久化 file→version 映射——version 令牌由模型持有，进程重启后旧令牌自然
# 失配→触发重读，正是我们要的安全行为。

def _file_version(target: Path) -> str:
    """文件内容的版本号 = sha256(raw bytes) 前 12 位 hex。对任意字节变换都鲁棒。"""
    return hashlib.sha256(target.read_bytes()).hexdigest()[:12]


def _check_version(target: Path, version: str) -> tuple[bool, str, str]:
    """校验文件是否仍是模型读取时的版本。
    返回 (ok, current_version, err_msg)：ok=False 时 err_msg 是给模型的提示。"""
    if not version:
        return False, "", ("[缺 version] 行号类编辑必须传 version——即你上次 "
                           "read_file/grep 返回的 file_version。先读文件拿到它。")
    current = _file_version(target)
    if current != version:
        return False, current, (
            f"[版本过期] 文件已改动（你的 version={version}，当前={current}），"
            f"行号可能已位移。请重新 read_file/grep 取最新行号与 file_version 后再编辑。")
    return True, current, ""


def _run_subprocess_streaming(args, name, shell=False, env=None):
    """运行子进程，实时流式输出 + 30 秒心跳进度。reader 线程兼容 Windows。
    通过 _tool_emit 回调推送 tool_stream / tool_progress 事件。
    env: 附加环境变量（None=继承父进程）；run_python 用它注入 PY_ARGS。
    Windows 加 CREATE_NO_WINDOW：detached（/restart 看门狗拉起）进程无控制台时，
    子进程默认各弹一个终端窗，闪退即此。"""
    proc = subprocess.Popen(
        args, stdin=subprocess.DEVNULL, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        text=True, cwd=str(WORKSPACE), shell=shell, env=env,
        creationflags=(subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0),
        bufsize=1, encoding="utf-8", errors="replace",
    )
    start = time.time()

    # reader 线程：逐行读 stdout → queue
    line_q: queue.Queue = queue.Queue()

    def _reader():
        try:
            for line in proc.stdout:
                line_q.put(line)
        except Exception:
            pass
        line_q.put(None)  # EOF

    threading.Thread(target=_reader, daemon=True).start()

    output_lines = []
    stream_buf = []
    last_hb = start
    last_flush = start

    while True:
        try:
            line = line_q.get(timeout=0.5)
        except queue.Empty:
            line = "__poll__"  # 无输出，走心跳/超时检查

        if line is None:
            break  # EOF
        elif line != "__poll__":
            output_lines.append(line)
            stream_buf.append(line)

        now = time.time()
        elapsed = now - start

        # 流式输出（每 ~1 秒 flush 一次，避免事件风暴）
        if stream_buf and now - last_flush >= 1.0:
            if _tool_emit:
                _tool_emit({"type": "tool_stream", "name": name,
                            "text": "".join(stream_buf), "elapsed": round(elapsed, 1)})
            stream_buf = []
            last_flush = now

        # 心跳（每 30 秒）
        if now - last_hb >= 30.0:
            if _tool_emit:
                preview = "".join(output_lines[-5:])[-200:]
                _tool_emit({"type": "tool_progress", "name": name,
                            "elapsed": round(elapsed, 1), "lines": len(output_lines),
                            "preview": preview})
            last_hb = now

        # 超时 → 转后台（不杀进程，daemon 线程继续读输出）
        if elapsed > TOOL_TIMEOUT:
            # flush 最后的流式输出
            if stream_buf and _tool_emit:
                _tool_emit({"type": "tool_stream", "name": name, "text": "".join(stream_buf)})
                stream_buf = []
            # 注册后台任务，daemon 线程继续读输出
            bg_id = f"bg_{int(time.time()*1000)}"
            _bg_tasks[bg_id] = {
                "name": name, "proc": proc, "output": output_lines[:],
                "started_at": start, "finished": False, "returncode": None,
            }
            def _bg_reader(_bg_id=bg_id, _line_q=line_q, _task=_bg_tasks[bg_id]):
                while True:
                    try:
                        ln = _line_q.get(timeout=1.0)
                    except queue.Empty:
                        if _task["proc"].poll() is not None:
                            break
                        continue
                    if ln is None:
                        break
                    _task["output"].append(ln)
                _task["proc"].wait()
                _task["returncode"] = _task["proc"].returncode
                _task["finished"] = True
                if _tool_emit:
                    _tool_emit({"type": "tool_stream", "name": name,
                                "text": f"\n[后台任务 {_bg_id} 已完成，返回码={_task['returncode']}]"})
            threading.Thread(target=_bg_reader, daemon=True).start()
            return (f"[执行超过 {TOOL_TIMEOUT}s，已转后台运行（任务ID: {bg_id}）。\n"
                    f"进程继续运行，不阻塞当前工作。用 check_bg_task(\"{bg_id}\") 查看进度和结果。"
                    f"已捕获 {len(output_lines)} 行输出。]")

    proc.wait()
    # 最终 flush
    if stream_buf and _tool_emit:
        _tool_emit({"type": "tool_stream", "name": name, "text": "".join(stream_buf)})

    return "".join(output_lines).strip() or "(无输出)"


def run_python(code: str = "", file: str = "", args: str = "") -> str:
    """运行 Python，实时流式输出（支持长任务进度）。独立子进程执行。二选一：
    - code：一段内联 Python 代码（写临时文件再跑）；
    - file：运行一个已存在的 .py 文件（跑已保存的脚本用这个，别再用 subprocess 包壳）。
    args: 传给脚本的参数字符串（子进程经环境变量 PY_ARGS 读取：
      `import os; a = os.environ.get("PY_ARGS", "")`——可放 JSON/CSV 等任意格式，脚本自行解析）。
      file 和 code 模式都生效，让已保存脚本可参数化复用（同类脚本不同参数不必改代码）。"""
    env = None
    if args:
        env = dict(os.environ)
        env["PY_ARGS"] = str(args)
    if file:
        target = _resolve(file)
        if not target.exists():
            return f"[文件不存在] {file}"
        return _run_subprocess_streaming([sys.executable, str(target)], f"run_python {file}", env=env)
    if not code:
        return "[参数缺失] run_python 需传 code（内联代码）或 file（.py 文件路径）"
    with tempfile.NamedTemporaryFile("w", suffix=".py", delete=False, encoding="utf-8") as f:
        f.write(code)
        tmp = f.name
    try:
        return _run_subprocess_streaming([sys.executable, tmp], "run_python", env=env)
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass


def _number_lines(text: str) -> str:
    """给一段文本的每行加行号（宽度按总行数自适应），格式 `N│ 内容`，与 read_file/find_function 同款。
    recent-file 快照用它：模型看到带行号的最新全文 + 标签里的 version，即可直接按行号 insert/delete/move。
    超 4000 行首尾各留 2000 行；尾部用【真实行号】编号，模型能看到尾段真实位置。"""
    lines = text.splitlines()
    n = len(lines)
    w = len(str(n))
    if n <= 4000:
        return "\n".join(f"{i:>{w}}│ {ln}" for i, ln in enumerate(lines, 1))
    head = "\n".join(f"{i:>{w}}│ {ln}" for i, ln in enumerate(lines[:2000], 1))
    trail = "\n".join(f"{i:>{w}}│ {ln}" for i, ln in enumerate(lines[-2000:], n - 1999))
    return head + "\n... (共{}行，需全文调 read_file)".format(n) + "\n" + trail


def _md_headings(text: str) -> list:
    """提取 Markdown 的 ATX 标题（#~######），返回 [(行号1based, 层级, 标题), ...]。
    跳过 ``` / ~~~ 代码围栏里的 #。供 _md_snapshot 结构目录、wiki_list/wiki_tree 大纲复用。"""
    in_fence, fence = False, ""
    out = []
    for idx, raw in enumerate(text.splitlines()):
        s = raw.strip()
        if s[:3] in ("```", "~~~"):
            if not in_fence:
                in_fence, fence = True, s[:3]
            elif s == fence:
                in_fence = False
            continue
        if in_fence:
            continue
        m = re.match(r"^(#{1,6})\s+(.*?)\s*#*$", raw)
        if m:
            out.append((idx + 1, len(m.group(1)), m.group(2).strip()))
    return out


def _md_outline(text: str) -> str:
    """Markdown 大纲：frontmatter + 各级标题 → 行范围树（缩进表层级，跳过代码围栏里的 #）。
    _md_snapshot 的 <structure> 部分独立抽出，两处共用；dir_outline 的 .md 文件展开也用它。"""
    lines = text.splitlines()
    n = len(lines)
    entries = []  # (start_1based, end_1based, label, level)  level=0 给 frontmatter
    frontmatter_end = 0   # 1-based：闭合 --- 所在行（0 = 无 frontmatter）
    if n >= 2 and lines[0].strip() == "---":
        close = next((j for j in range(1, n) if lines[j].strip() == "---"), None)
        if close is not None:
            entries.append((1, close + 1, "frontmatter", 0))
            frontmatter_end = close + 1
    # 标题（排除 frontmatter 区域，避免把 YAML 注释 # 当标题）
    headings = [(ln, lv, t) for (ln, lv, t) in _md_headings(text) if ln > frontmatter_end]
    for k, (ln, level, title) in enumerate(headings):
        nxt = next((ln2 for (ln2, lv2, _) in headings[k + 1:] if lv2 <= level), None)
        end = (nxt - 1) if nxt is not None else n   # 到下一个同级/更高级标题前
        entries.append((ln, end, title, level))
    struct_lines = []
    for a, b, label, level in entries:
        indent = "  " * (level - 1)
        rng = f"[L{a}-L{b}]" if a != b else f"[L{a}]"
        struct_lines.append(f"{indent}{rng} {label}")
    return "\n".join(struct_lines) or "(无 frontmatter / 标题)"


def _md_snapshot(text: str) -> str:
    """Markdown 快照：<structure> 结构目录（frontmatter + ATX 标题 → 行范围，缩进表层级，跳过代码
    围栏里的 #）+ <content> 干净正文（不带 N│ 行号——结构目录取代行号做 .md 的导航）。
    recent-file 和 read_file(.md) 都用它。超 4000 行时正文首尾截断、结构目录保持完整。"""
    struct = _md_outline(text)
    n = len(text.splitlines())
    # 正文（超长首尾截断，结构目录保持完整）
    if n <= 4000:
        body = text.rstrip("\n")
    else:
        lines = text.splitlines()
        body = ("\n".join(lines[:2000]) + f"\n... (共{n}行，需全文调 read_file)\n"
                + "\n".join(lines[-2000:])).rstrip("\n")
    return f"<structure>\n{struct}\n</structure>\n<content>\n{body}\n</content>"


def read_file(path: str, start_line: int = None, end_line: int = None,
              line_numbers: bool = True) -> str:
    """读取 workspace 内某个文件的内容（统一入口）：文本/Word/Excel/PDF 自动提取、
    图片（png/jpg/gif/webp/bmp）自动转 data URL（视觉模型可直接查看），任何文件都
    只用 read_file。末尾附 file_version（图片除外）。
    start_line/end_line: 只读指定行范围（1-based，含两端；不传=全文）。
    line_numbers: 默认 True，每行前加行号（宽度按本段最大行号自适应对齐），用于接下来要用
    insert/delete/move 按行号编辑的场景；传 False 得不含行号的纯文本。
    对 .md 文件特例：line_numbers=True 时返回 <structure> 结构目录 + <content> 干净正文
    （frontmatter/标题→行范围，结构取代行号做导航）；line_numbers=False 仍是纯文本。
    返回末尾的 file_version 是该文件当前的内容版本号——传给 insert/delete/move 的 version 参数；
    若编辑时版本对不上，说明文件已被改动、需重读。"""
    target = _resolve(path)
    # 图片类型 → 走读图逻辑（data URL，视觉模型可查看）。放在存在性检查之前：
    # _read_image 自带候选查找（cwd + repo images/，供 <img> 标签裸文件名回读），
    # 找不到时给 [未找到图片]；非视觉模型调用也只会得到 <img> 占位提示。
    if target.suffix.lower() in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}:
        return _read_image(path)
    if not target.exists():
        return f"[文件不存在] {path}"
    if target.suffix.lower() in {".docx", ".xlsx", ".xlsm", ".xltx", ".pdf"}:
        text = _extract_text(target)
        if text is None:
            text = target.read_text(encoding="utf-8", errors="ignore")
        lines = text.splitlines()
    else:
        text = target.read_text(encoding="utf-8")
        lines = text.splitlines()
    total = len(lines)
    ver_footer = f"\n[file_version={_file_version(target)}]"
    _EXCEL_EXTS = {".xlsx", ".xlsm", ".xltx"}
    if start_line is None and end_line is None:
        if line_numbers and target.suffix.lower() in {".md", ".markdown"}:
            return _md_snapshot(text) + ver_footer
        if target.suffix.lower() in _EXCEL_EXTS:
            # Excel 提取文本已内嵌【Sheet 内 1-based 行号】（=== Sheet: xxx === 分节），
            # 不再叠全局行号（双重行号反而干扰）；分页 start/end_line 仍按全局行计
            nsheets = text.count("=== Sheet:")
            return (f"[{path} · Excel · {nsheets} 个 Sheet · 提取 {total} 行"
                    f"（行号为各 Sheet 内 1-based，与 Excel UI 一致）]\n{text}{ver_footer}")
        if line_numbers:
            w = len(str(total))
            body = "\n".join(f"{i+1:>{w}}│ {ln}" for i, ln in enumerate(lines))
            return f"[{path} 共 {total} 行]\n{body}{ver_footer}"
        return text + ver_footer
    start = max(1, start_line or 1) - 1
    end = min(total, end_line or total)
    if start >= total:
        return f"[行号越界] 文件共 {total} 行，请求 start_line={start_line}"
    selected = lines[start:end]
    header = f"[{path} L{start+1}-L{end}/{total}]"
    if line_numbers:
        w = len(str(end))
        body = "\n".join(f"{start+i+1:>{w}}│ {ln}" for i, ln in enumerate(selected))
    else:
        body = "\n".join(selected)
    return header + "\n" + body + ver_footer


# 视觉模型图片输入边长上限（qwen/GLM 等多为 2048，取保守值；超限会被 API 400 拒绝）
_MAX_IMG_EDGE = 2048


def _shrink_image_bytes(raw: bytes) -> tuple[bytes, str]:
    """图片任一边超过 _MAX_IMG_EDGE 时等比缩小到限内（需 Pillow；不可用/失败则原样返回）。
    返回 (bytes, mime)：mime 非空表示已转码（统一存 PNG，无损保文字清晰度；GIF 取首帧）。"""
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(raw))
        w, h = img.size
        if w <= _MAX_IMG_EDGE and h <= _MAX_IMG_EDGE:
            return raw, ""
        ratio = min(_MAX_IMG_EDGE / w, _MAX_IMG_EDGE / h)
        img = img.convert("RGBA" if img.mode in ("P", "LA") else img.mode) \
            .resize((max(1, int(w * ratio)), max(1, int(h * ratio))), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue(), "image/png"
    except Exception:
        return raw, ""


def _read_image(file: str) -> str:
    """（内部函数，由 read_file 统一入口调用，不单独注册为工具）读图返回 data URL。
    file 可以是裸文件名（如 c7_0.png：先在 cwd 找，再在 repo images/ 目录找），
    或相对/绝对路径（在 cwd 下解析，沙箱限定）。支持 png/jpg/gif/webp。
    超过 2048×2048 的图自动等比压缩到限内（视觉模型 API 的尺寸上限，超限会被拒）。
    仅视觉模型能查看返回的图；非视觉模型调用了也只会得到 <img> 占位。"""
    name = (file or "").strip().strip('"').strip("'")
    if not name:
        return "[未传文件名]"
    cands = []
    if os.sep in name or "/" in name or "\\" in name:   # 带路径：workspace 沙箱解析
        try:
            cands.append(_resolve(name))
        except Exception:
            pass
        cands.append(WORKSPACE / name)
    else:                                                # 裸文件名：cwd + repo images/
        cands.append(WORKSPACE / name)
        try:
            from session import repo_images_dir
            cands.append(repo_images_dir(WORKSPACE) / name)
        except Exception:
            pass
    for p in cands:
        try:
            if p.exists() and p.is_file():
                ext = (p.suffix.lstrip(".") or "png").lower()
                ext = "jpeg" if ext == "jpg" else ext
                raw = p.read_bytes()
                raw, forced_mime = _shrink_image_bytes(raw)   # 超限自动压缩
                mime = forced_mime or mimetypes.guess_type(str(p))[0] or f"image/{ext}"
                b64 = base64.b64encode(raw).decode()
                return f"data:{mime};base64,{b64}"
        except Exception:
            continue
    return f"[未找到图片] {file}（cwd 和 repo images/ 都没找到）"


def write_file(path: str, content: str) -> str:
    """把 content 写入 workspace 内的文件（覆盖），返回确认信息。"""
    target = _resolve(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(content, encoding="utf-8")
    return f"已写入 {len(content)} 字符到 {path}"


def list_dir(path: str = ".") -> str:
    """列出 workspace 内某目录下的文件/子目录。"""
    target = _resolve(path)
    if not target.exists():
        return f"[目录不存在] {path}"
    entries = sorted(p.name + ("/" if p.is_dir() else "") for p in target.iterdir())
    return "\n".join(entries) if entries else "(空目录)"


def _extract_text(target: Path) -> str | None:
    """对 Word/Excel/PDF 提取纯文本；不支持则返回 None。"""
    suffix = target.suffix.lower()
    try:
        if suffix == ".docx":
            import docx
            doc = docx.Document(str(target))
            return "\n".join(p.text for p in doc.paragraphs)
        if suffix in (".xlsx", ".xlsm", ".xltx"):
            import openpyxl
            wb = openpyxl.load_workbook(str(target), read_only=True, data_only=True)
            parts = []
            for name in wb.sheetnames:
                ws = wb[name]
                parts.append(f"=== Sheet: {name} ===")
                w = len(str(ws.max_row or 1))
                for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                    parts.append(f"{i:>{w}}│ " + "\t".join(str(c) if c is not None else "" for c in row))
            wb.close()
            return "\n".join(parts)
        if suffix == ".pdf":
            text = ""
            try:
                import fitz
                doc = fitz.open(str(target))
                for page in doc:
                    text += page.get_text()
                doc.close()
            except Exception:
                import PyPDF2
                reader = PyPDF2.PdfReader(str(target))
                for page in reader.pages:
                    t = page.extract_text()
                    if t:
                        text += t + "\n"
            return text.strip()
        return None
    except Exception as e:
        return f"[文档解析失败: {type(e).__name__}: {e}]"


def grep(pattern: str, path: str = ".", glob: str = None, regex: bool = True,
         context: int = 0, max_results: int = 50) -> str:
    """在 workspace 内搜索文件内容，返回带行号的匹配（可带上下文），并附每个文件的 file_version。
    pattern: 搜索模式，**默认按正则**（支持 a|b 多选一、. * 等元字符，与 ripgrep 一致）；
    要按字面匹配（把 pattern 当普通字符串）传 regex=False。path: 文件或目录(默认 workspace 根)。
    glob: 文件名过滤如 '*.js'；context: 每条命中前后各显示几行（默认 0=只显示匹配行）；
    max_results: 最多返回匹配数。每个文件头部的 file_version 传给 insert/delete/move 的 version 参数。"""
    import fnmatch
    import re
    root = _resolve(path)
    if not root.exists():
        return f"[路径不存在] {path}"
    try:
        rx = re.compile(pattern if regex else re.escape(pattern))
    except re.error as e:
        return f"[正则错误] {e}\n（pattern 含特殊字符？可传 regex=False 按字面匹配）"
    DOC_EXT = {".docx", ".xlsx", ".xlsm", ".xltx", ".pdf"}
    # path 是文件 → 只搜该文件；是目录 → 递归其下（rglob 在文件上返回空，故需分支）
    candidates = [root] if root.is_file() else sorted(root.rglob("*"))
    # 按文件聚合：rel -> (version, lines, [命中行号])
    files, scanned, total = {}, 0, 0
    capped = False
    for fp in candidates:
        if not fp.is_file() or (glob and not fnmatch.fnmatch(fp.name, glob)):
            continue
        text = None
        if fp.suffix.lower() in DOC_EXT:
            extracted = _extract_text(fp)
            if extracted and not extracted.startswith("[文档解析失败"):
                text = extracted
        else:
            try:
                text = fp.read_text(encoding="utf-8", errors="ignore")
            except Exception:
                continue
        if text is None:
            continue
        scanned += 1
        lines = text.splitlines()
        hits = []
        for i, line in enumerate(lines, 1):
            if rx.search(line):
                hits.append(i)
                total += 1
                if total >= max_results:
                    capped = True
                    break
        if hits:
            rel = fp.relative_to(WORKSPACE).as_posix()
            files[rel] = (_file_version(fp), lines, hits)
        if capped:
            break
    if not files:
        return f"(扫描 {scanned} 个文件，未找到 '{pattern}')"
    c = max(0, int(context))
    parts = [f"扫描 {scanned} 个文件，匹配 {total} 处："]
    for rel, (ver, lines, hits) in files.items():
        n = len(lines)
        parts.append(f"── {rel}  (file_version={ver}) ──")
        for lineno in hits:
            if c:
                lo, hi = max(1, lineno - c), min(n, lineno + c)
                for j in range(lo, hi + 1):
                    mark = ">" if j == lineno else " "
                    parts.append(f"{mark} {rel}:{j}: {lines[j-1].rstrip()[:200]}")
            else:
                parts.append(f"> {rel}:{lineno}: {lines[lineno-1].rstrip()[:200]}")
    if capped:
        parts.append(f"...（已达 max_results={max_results}，截断；收紧 pattern 或调大 max_results）")
    return "\n".join(parts)


def edit(path: str, old_string: str, new_string: str, replace_all: bool = False,
         start_line: int = None, end_line: int = None) -> str:
    """精确替换文件中的一段文本。
    path: workspace 内文件；old_string: 要替换的原文；new_string: 替换为；
    replace_all=True 替换全部匹配。
    start_line/end_line: 只在该行范围内搜索替换（1-based，含两端）。"""
    target = _resolve(path)
    if not target.exists():
        return f"[文件不存在] {path}"
    content = target.read_text(encoding="utf-8")
    lines = content.splitlines()
    total = len(lines)
    # 行范围限定
    if start_line is not None or end_line is not None:
        s = max(0, (start_line or 1) - 1)
        e = min(total, end_line or total)
        if s >= total:
            return f"[行号越界] 文件共 {total} 行，start_line={start_line}"
        scope = "\n".join(lines[s:e])
        prefix = "\n".join(lines[:s])
        suffix = "\n".join(lines[e:])
    else:
        scope = content
        prefix, suffix = "", ""
        s = 0
    count = scope.count(old_string)
    if count == 0:
        # 精确匹配失败 → 回退：去每行【行尾】空白后做行级比对（唯一则接受；多处则不唯一）。
        # 只去行尾、不碰行首（缩进是 Python 语义）；tab↔空格不自动规整（tabstop 未知、易跨层级误配）。
        scope_lines = lines[s:e] if (start_line or end_line) else lines
        old_lines = old_string.splitlines()
        m = len(old_lines)
        fp = [ln.rstrip() for ln in scope_lines]
        ofp = [ln.rstrip() for ln in old_lines]
        starts = [i for i in range(len(fp) - m + 1) if fp[i:i + m] == ofp] if (m and m <= len(fp)) else []
        if not starts:
            where = f" L{s+1}-L{min(e,total) if (start_line or end_line) else total}" if (start_line or end_line) else ""
            return (f"[未找到]{where} 精确与去行尾空白后均未命中 old_string；"
                    f"常见原因：行尾空格、缩进 tab↔空格、或内容已改——重新 read_file 取准确文本")
        if len(starts) > 1 and not replace_all:
            return f"[不唯一] 去行尾空白后共匹配 {len(starts)} 处，请加更多上下文让 old_string 唯一，或设 replace_all=True"
        hits = starts if replace_all else starts[:1]
        repl = new_string.splitlines()
        for i in sorted(hits, reverse=True):
            scope_lines[i:i + m] = repl
        new_scope = "\n".join(scope_lines)
        if not (start_line or end_line) and content.endswith("\n") and not new_scope.endswith("\n"):
            new_scope += "\n"   # splitlines+join 吃掉了文末换行，补回（精确路径本就保留）
        new_content = (prefix + ("\n" if prefix else "") + new_scope + ("\n" if suffix else "") + suffix) if (start_line or end_line) else new_scope
        new_content = (prefix + ("\n" if prefix else "") + new_scope + ("\n" if suffix else "") + suffix) if (start_line or end_line) else new_scope
        target.write_text(new_content, encoding="utf-8")
        return f"✅ 已替换 {len(hits)} 处（行尾空白容忍匹配，{path}" + (f" L{start_line}-L{end_line}" if start_line or end_line else "") + ")"
    if count > 1 and not replace_all:
        return f"[不唯一] 共匹配 {count} 处，请加更多上下文让 old_string 唯一，或设 replace_all=True"
    if old_string == new_string:
        return "[无变化] old_string 与 new_string 相同"
    new_scope = scope.replace(old_string, new_string) if replace_all else scope.replace(old_string, new_string, 1)
    new_content = (prefix + ("\n" if prefix else "") + new_scope + ("\n" if suffix else "") + suffix) if (start_line or end_line) else new_scope
    new_content = (prefix + ("\n" if prefix else "") + new_scope + ("\n" if suffix else "") + suffix) if (start_line or end_line) else new_scope
    target.write_text(new_content, encoding="utf-8")
    return f"✅ 已替换 {count if replace_all else 1} 处（{path}" + (f" L{start_line}-L{end_line}" if start_line or end_line else "") + ")"


def _apply_lines(target: Path, new_lines: list, path: str, action_desc: str) -> str:
    """把 new_lines 写回文件（统一 \n 换行 + 末尾换行），返回带新 file_version 的确认串。"""
    target.write_text("\n".join(new_lines) + ("\n" if new_lines else ""), encoding="utf-8")
    return f"{action_desc}（现共 {len(new_lines)} 行）file_version={_file_version(target)}"


def insert(path: str, entries: list, version: str) -> str:
    """按行号在文件中【一处或多处】插入文本，单次原子写入——一次插多段用它，别在 run_python 里拼字符串。
    entries: 插入点数组，每项 {"line": 1-based 行号, "content": 文本(可多行)}；在该行之前插入；
             line <=0 或超过总行数则追加末尾。
    内部先按 line 排序、再【降序】应用（先插高位不扰动低位行号），故直接传 grep/read_file 查到的原始行号即可，无需自己算位移。
    需传 read_file/grep 返回的 file_version 校验（不匹配=文件已改、拒绝要求重读）；成功返回新 file_version。"""
    target = _resolve(path)
    if not target.exists():
        return f"[文件不存在] {path}"
    # 校验 entries
    if not isinstance(entries, list) or not entries:
        return f"[参数错误] entries 需为非空数组，收到 {type(entries).__name__}"
    norm = []
    for i, e in enumerate(entries):
        if not isinstance(e, dict):
            return f"[参数错误] entries[{i}] 需为对象 {{line, content}}，收到 {type(e).__name__}"
        ln = e.get("line")
        ct = e.get("content")
        if not isinstance(ln, int) or isinstance(ln, bool):
            return f"[参数错误] entries[{i}].line 需为整数，收到 {ln!r}"
        if not isinstance(ct, str):
            return f"[参数错误] entries[{i}].content 需为字符串，收到 {type(ct).__name__}"
        norm.append((ln, ct))
    ok, _cur, err = _check_version(target, version)
    if not ok:
        return err
    out = target.read_text(encoding="utf-8").splitlines()
    total = len(out)
    # 先按 line 排序（升序），再降序应用：先插大行号，不影响小行号位置
    norm.sort(key=lambda ec: ec[0], reverse=True)
    appended = 0
    for ln, ct in norm:
        block = (ct or "").splitlines()
        if ln <= 0 or ln > total:
            out.extend(block)             # 追加到末尾
            appended += 1
        else:
            out[ln - 1:ln - 1] = block    # 在第 ln 行之前插入
    ins_n = len(norm) - appended
    where = f"{ins_n} 处定点" + (f"+ {appended} 处追加" if appended else "")
    total_new_lines = sum(len((c or "").splitlines()) for _, c in norm)
    return "✅ " + _apply_lines(target, out, path,
                                f"已在 {path} 插入 {where}（共 {total_new_lines} 行）")


def delete(path: str, start_line: int, end_line: int, version: str) -> str:
    """按行号删除文件中一段连续的行（start_line~end_line，含两端，1-based）。删的是行不是文件。
    需传 read_file/grep 返回的 file_version 做版本校验：不匹配则文件已改动、会被拒绝要求重读。
    成功后返回新的 file_version，同轮后续编辑可直接用它当 version。"""
    target = _resolve(path)
    if not target.exists():
        return f"[文件不存在] {path}"
    ok, _cur, err = _check_version(target, version)
    if not ok:
        return err
    lines = target.read_text(encoding="utf-8").splitlines()
    total = len(lines)
    s = int(start_line)
    e = int(end_line)
    if s < 1 or s > total:
        return f"[行号越界] 文件共 {total} 行，start_line={start_line}（须 1~{total}）"
    if e < s:
        return f"[参数错误] end_line({end_line}) 不能小于 start_line({start_line})"
    e = min(total, e)   # 超出末尾则截到文件尾（友善：删到末尾）
    del lines[s - 1:e]
    return "✅ " + _apply_lines(target, lines, path,
                                f"已删除 {path} 第 {s}-{e} 行（共 {e - s + 1} 行）")


def move(path: str, start_line: int, end_line: int, dst_line: int, version: str) -> str:
    """按行号把一段行（start_line~end_line，含两端）原子搬到 dst_line 行之前——重构搬代码块用。
    需传 read_file/grep 返回的 file_version 做版本校验：不匹配则文件已改动、会被拒绝要求重读。
    dst_line 按原始行号理解（搬到自身范围内视为无操作）。成功后返回新的 file_version。"""
    target = _resolve(path)
    if not target.exists():
        return f"[文件不存在] {path}"
    ok, _cur, err = _check_version(target, version)
    if not ok:
        return err
    lines = target.read_text(encoding="utf-8").splitlines()
    total = len(lines)
    s = int(start_line)
    e = int(end_line)
    if s < 1 or s > total:
        return f"[行号越界] 文件共 {total} 行，start_line={start_line}（须 1~{total}）"
    if e < s:
        return f"[参数错误] end_line({end_line}) 不能小于 start_line({start_line})"
    e = min(total, e)
    block = lines[s - 1:e]
    nblock = len(block)
    d = int(dst_line)
    # 目标落在源块自身范围内 → 无操作（否则行号语义自相矛盾）
    if s <= d <= e + 1:
        return f"[无操作] dst_line={d} 落在源块 {s}-{e} 内，无需移动。file_version={_file_version(target)}"
    remaining = lines[:s - 1] + lines[e:]
    # 把原始行号语义映射到 remaining 的插入下标
    if d <= s:
        ins = max(0, d - 1)
    else:  # d > e：源块已在 dst 之前被整体移除，dst 在 remaining 中下标前移 nblock
        ins = max(0, d - 1 - nblock)
    ins = min(ins, len(remaining))
    remaining[ins:ins] = block
    return "✅ " + _apply_lines(target, remaining, path,
                                f"已把 {path} 第 {s}-{e} 行（{nblock} 行）搬到原第 {d} 行前")


def replace_lines(path: str, entries: list, version: str) -> str:
    """按行号【一处或多处】整段替换文件内容，单次原子写入——重写整个函数/大段代码用它（比 edit 省 token，不必重吐旧文本）。
    entries: 替换段数组，每项 {"range": [起, 止], "content": 新文本(可多行)}；range 1-based 含两端；
             [n,n] 替换第 n 行；content="" 删除该范围（等价 delete）。多处直接传 read_file/grep 查到的原始行号即可——
             内部按 range 起点降序应用（先改高位不扰动低位行号），各段 range 不许重叠。
    需传 read_file/grep 返回的 file_version 校验（不匹配=文件已改、拒绝要求重读）；成功返回新 file_version。"""
    target = _resolve(path)
    if not target.exists():
        return f"[文件不存在] {path}"
    if not isinstance(entries, list) or not entries:
        return f"[参数错误] entries 需为非空数组，收到 {type(entries).__name__}"
    norm = []
    for i, e in enumerate(entries):
        if not isinstance(e, dict):
            return f"[参数错误] entries[{i}] 需为对象 {{range, content}}，收到 {type(e).__name__}"
        rng = e.get("range")
        ct = e.get("content")
        if not (isinstance(rng, list) and len(rng) == 2
                and all(isinstance(x, int) and not isinstance(x, bool) for x in rng)):
            return f"[参数错误] entries[{i}].range 需为两个整数 [起, 止]，收到 {rng!r}"
        if not isinstance(ct, str):
            return f"[参数错误] entries[{i}].content 需为字符串，收到 {type(ct).__name__}"
        a, b = rng
        if a < 1 or a > b:
            return f"[参数错误] entries[{i}].range={rng} 非法：须 1≤起≤止"
        norm.append([a, b, ct])
    ok, _cur, err = _check_version(target, version)
    if not ok:
        return err
    lines = target.read_text(encoding="utf-8").splitlines()
    total = len(lines)
    for a, _b, _ in norm:
        if a > total:
            return f"[行号越界] 文件共 {total} 行，range 起={a}（须 ≤{total}）"
    for seg in norm:                 # 止点截到文件尾（友善，同 delete）
        seg[1] = min(total, seg[1])
    asc = sorted(norm, key=lambda x: x[0])   # 非重叠校验：升序看相邻段是否相交
    for (a1, b1, _), (a2, _b2, _) in zip(asc, asc[1:]):
        if a2 <= b1:
            return f"[参数错误] range 重叠：[{a1},{b1}] 与起点 {a2} 的段相交，请合并或调整行号"
    for a, b, ct in sorted(norm, key=lambda x: x[0], reverse=True):   # 降序：先改高位
        lines[a - 1:b] = (ct or "").splitlines()
    new_lines = sum(len((ct or "").splitlines()) for _a, _b, ct in norm)
    return "✅ " + _apply_lines(target, lines, path,
                                f"已在 {path} 整段替换 {len(norm)} 处（新内容共 {new_lines} 行）")


# ===== 函数定位（find_function）=====
# 比 grep 更适合"看某个函数的完整实现"：定位签名起始行后，按缩进(Python)或大括号配对
# (其它语言)找到函数结束行，整段带行号返回。大括号配对用字符级状态机，跳过字符串/注释，
# 避免把 "}" 或 // } 里的括号算进去。

_LANG_FAMILY = {
    ".py": "python",
    ".js": "brace", ".jsx": "brace", ".mjs": "brace", ".cjs": "brace",
    ".ts": "brace", ".tsx": "brace",
    ".cs": "brace", ".java": "brace", ".kt": "brace",
    ".c": "brace", ".h": "brace", ".cpp": "brace", ".cc": "brace",
    ".cxx": "brace", ".hpp": "brace", ".hxx": "brace", ".ino": "brace",
    ".go": "brace", ".rs": "brace", ".swift": "brace", ".php": "brace", ".scala": "brace",
}


def _lang_family(target: Path, lang: str = None) -> str:
    """返回 'python' / 'brace' / None。lang 显式指定时优先；否则按扩展名识别。"""
    if lang:
        return "python" if lang.lower() in ("py", "python") else "brace"
    return _LANG_FAMILY.get(target.suffix.lower())


def _indent(line: str) -> int:
    """行首空白（空格/Tab）的字符数。"""
    return len(line) - len(line.lstrip(" \t"))


def _extend_decorators(lines: list, idx: int) -> int:
    """Python：把 def 上方连续的、同缩进的 @装饰器 行并入起始。"""
    sig = _indent(lines[idx])
    j = idx - 1
    while j >= 0 and _indent(lines[j]) == sig and lines[j].lstrip().startswith("@"):
        j -= 1
    return j + 1


def _find_def_starts(lines: list, name: str, family: str) -> list:
    """返回所有疑似 NAME 定义的起始行 index（未确认是否有体）。"""
    import re
    nm = re.escape(name)
    starts = []
    if family == "python":
        rx = re.compile(rf"^[ \t]*(?:async[ \t]+)?def[ \t]+{nm}\b")
        starts = [i for i, ln in enumerate(lines) if rx.match(ln)]
    else:  # 大括号家族
        form_call = re.compile(rf"\b{nm}\s*(?:<[^>]*>)?\s*\(")   # 含简单泛型 <T>
        for i, ln in enumerate(lines):
            hit = False
            m = form_call.search(ln)
            if m:
                pre = ln[:m.start()].rstrip()
                if not (pre and pre[-1] == "."):   # 排除 obj.method( 这类调用
                    hit = True
            if not hit and re.search(rf"\b{nm}\s*=", ln) and \
                    ("=>" in ln or re.search(r"=\s*(?:async\s*)?function\b", ln)):
                hit = True   # const NAME = (...) =>  /  NAME = function
            if hit:
                starts.append(i)
    return starts


def _find_block_end_brace(lines: list, start_idx: int):
    """大括号家族：从 start_idx 扫，首个 { 开体、配对到其闭合 } 为止，返回结束行 index。
    扫描跳过字符串/注释；若开体前先撞到 ;（声明/调用，无体）返回 None。"""
    depth = 0
    opened = False
    in_str = None        # 当前字符串引号，或 None
    in_line = False      # // 行注释
    in_block = False     # /* */ 块注释
    i = start_idx
    while i < len(lines):
        line = lines[i]
        j, n = 0, len(line)
        while j < n:
            c = line[j]
            nxt = line[j + 1] if j + 1 < n else ""
            if in_line:
                break
            if in_block:
                if c == "*" and nxt == "/":
                    in_block = False
                    j += 2
                    continue
                j += 1
                continue
            if in_str is not None:
                if c == "\\":
                    j += 2
                    continue
                if c == in_str:
                    in_str = None
                j += 1
                continue
            if c == "/" and nxt == "/":
                in_line = True
                break
            if c == "/" and nxt == "*":
                in_block = True
                j += 2
                continue
            if c in ('"', "'", "`"):
                in_str = c
                j += 1
                continue
            if c == "{":
                depth += 1
                opened = True
            elif c == "}":
                if opened:
                    depth -= 1
                    if depth == 0:
                        return i
            elif c == ";":
                if not opened:
                    return None   # 开体前遇 ; → 声明/调用，无函数体
            j += 1
        in_line = False
        i += 1
    return None


def _find_block_end_indent(lines: list, start_idx: int) -> int:
    """Python：def 行缩进为基准；体 = 缩进更深的行(空行暂属体)，遇到缩进 ≤ 基准的非空行结束。
    末尾去掉所属的空行。返回结束行 index。"""
    sig = _indent(lines[start_idx])
    end = start_idx
    for i in range(start_idx + 1, len(lines)):
        ln = lines[i]
        if ln.strip() == "" or _indent(ln) > sig:
            end = i
        else:
            break
    while end > start_idx and lines[end].strip() == "":
        end -= 1
    return end


def find_function(name: str, path: str, lang: str = None, context: int = 0) -> str:
    """查找某个函数/方法的完整定义，返回带行号的函数体 + 行范围 + file_version（比 grep 更适合"看某函数完整实现"）。
    name: 函数/方法名（精确，非正则）；path: 文件或目录（目录则按语言扩展名扫描各文件，跨文件找定义）；
    lang: python/js/ts/cs/java/cpp/go...，不传则按扩展名识别；context: 函数体前后额外显示几行（默认0）。
    按 缩进(Python) / 大括号配对(其它) 自动定位起止行；返回的 [path L起-L止/总] 与 file_version 可直接喂给 insert/delete/move。"""
    target = _resolve(path)
    if not target.exists():
        return f"[路径不存在] {path}"
    if target.is_dir():
        exts = set(_LANG_FAMILY.keys())
        files = [p for p in sorted(target.rglob("*")) if p.is_file() and p.suffix.lower() in exts]
    else:
        files = [target]
    c = max(0, int(context))
    parts = []
    total_matches = 0
    for fp in files:
        family = _lang_family(fp, lang)
        if not family:
            continue
        try:
            text = fp.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        lines = text.splitlines()
        total = len(lines)
        blocks = []
        for s in _find_def_starts(lines, name, family):
            if family == "python":
                real_s = _extend_decorators(lines, s)
                e = _find_block_end_indent(lines, s)
            else:
                e = _find_block_end_brace(lines, s)
                if e is None:
                    continue   # 声明/调用，无体 → 跳过
                real_s = s
            blocks.append((real_s, e))
        if not blocks:
            continue
        rel = path if fp == target else fp.relative_to(WORKSPACE).as_posix()
        ver = _file_version(fp)
        for s, e in blocks:
            total_matches += 1
            lo, hi = max(0, s - c), min(total - 1, e + c)
            parts.append(f"[{rel} L{s + 1}-L{e + 1}/{total}]  file_version={ver}")
            w = len(str(hi + 1))
            for j in range(lo, hi + 1):
                parts.append(f"{j + 1:>{w}}│ {lines[j]}")
    if not parts:
        where = f"{path} 下" if target.is_dir() else f"{path} 中"
        return (f"(在{where}未找到 '{name}' 的函数定义)\n"
                f"可能：名字拼错 / 是无{{}}的表达式体箭头函数 / 仅声明无体 / 不支持的语言。可用 grep 按名字搜。")
    return f"找到 {total_matches} 处 '{name}' 的定义：\n" + "\n".join(parts)


# ===== 目录大纲（assembly DSL 的 dir: 装配项）=====

# brace 家族：显式定义关键字（class/interface/struct/enum/func/fn/function...）
_BRACE_DEF_HEAD = re.compile(r"^\s*(?:export\s+|default\s+|public\s+|private\s+|internal\s+|"
                             r"protected\s+|static\s+|abstract\s+|final\s+|sealed\s+|override\s+|"
                             r"partial\s+|async\s+)*(class|interface|struct|enum|trait|record|func|fn|function)\b")
# brace 家族：控制流/调用排除词——签名名或其前置词命中即不当作定义
_BRACE_CTRL_KW = frozenset({"if", "for", "while", "switch", "catch", "return", "using", "lock",
                            "foreach", "else", "do", "throw", "new", "await", "yield", "case",
                            "when", "match", "typeof", "sizeof", "delete", "in", "of"})
_PY_DEF_RX = re.compile(r"^(\s*)(?:async\s+)?(class|def)\s+([A-Za-z_]\w*)")
_BRACE_SIG_RX = re.compile(r"^(\s*)(?:[\w<>\[\]:,.*&~?]+\s+)*?([A-Za-z_]\w*)\s*(?:<[^>]*>)?\s*\(")
# 箭头/函数表达式赋值：const f = (x) => ... / NAME = function / let g = async x =>
_BRACE_ARROW_RX = re.compile(r"^\s*(?:export\s+|default\s+)?(?:const|let|var)\s+"
                             r"([A-Za-z_]\w*)\s*=\s*(?:async\s+)?"
                             r"(?:function\b|\([^)]*\)\s*=>|[A-Za-z_]\w*\s*=>)")


def _code_outline(text: str, family: str, max_defs: int = 40) -> str:
    """代码文件大纲：类/函数定义行号清单（dir_outline 展开 .py/.js/.cs 等用）。
    python = class / (async) def（缩进映射层级）；brace 家族 = 显式定义关键字行 +
    启发式方法签名行（排除控制流 / obj.method( 调用 / 赋值右值调用，箭头函数保留）。
    轻量正则不验证函数体——导航定位够用，精确边界用 find_function。"""
    out = []
    lines = text.splitlines()
    if family == "python":
        for i, ln in enumerate(lines):
            m = _PY_DEF_RX.match(ln)
            if m:
                ind = "  " * min(len(m.group(1)) // 2, 6)
                out.append(f"{ind}[L{i+1}] {m.group(2)} {m.group(3)}")
    else:
        for i, ln in enumerate(lines):
            s = ln.rstrip()
            st = s.strip()
            if not st or st.startswith(("//", "/*", "*", "#", "*")):
                continue
            m = _BRACE_DEF_HEAD.match(s)
            if m:
                out.append(f"  [L{i+1}] {_code_outline_trunc(st)}")
                continue
            if _BRACE_ARROW_RX.match(s):
                out.append(f"  [L{i+1}] {_code_outline_trunc(st)}")
                continue
            ms = _BRACE_SIG_RX.match(s)
            if not ms or ms.group(2) in _BRACE_CTRL_KW:
                continue
            pre = s[:ms.start(2)].rstrip()
            if pre and pre[-1] == ".":
                continue                       # obj.method( 方法调用
            pm = re.search(r"([A-Za-z_]\w*)\s*$", pre)
            if pm and pm.group(1) in _BRACE_CTRL_KW:
                continue                       # new Foo( / return Foo( / await fetch(
            head = s.split("(", 1)[0]
            if "=" in head and "=>" not in s:
                continue                       # x = foo( 赋值右值调用（const f = () => 保留）
            out.append(f"  [L{i+1}] {_code_outline_trunc(st)}")
    if len(out) > max_defs:
        out = out[:max_defs] + [f"  …另有 {len(out) - max_defs} 个定义（read_file 看全文）"]
    return "\n".join(out)


def _code_outline_trunc(s: str, w: int = 90) -> str:
    """大纲行截断到 w 字符（签名太长时保类名/函数名头部）。"""
    return s if len(s) <= w else s[:w - 1] + "…"


def _file_outline(fp: Path, pad: str) -> tuple[int, list[str]]:
    """(总行数, 大纲行)：.md → 标题行号树；代码文件 → 定义行号清单；其余只行数。
    超 512KB 不展开大纲（防爆 token），行数按字节流数。"""
    try:
        if fp.stat().st_size > 512_000:
            with fp.open("rb") as f:
                n = sum(buf.count(b"\n") for buf in iter(lambda: f.read(1 << 20), b""))
            return n, []
        text = fp.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return 0, []
    n = len(text.splitlines())
    suf = fp.suffix.lower()
    if suf in (".md", ".markdown"):
        o = _md_outline(text)
        if o and not o.startswith("(无"):
            return n, [f"{pad}{ln}" for ln in o.splitlines()]
    fam = _LANG_FAMILY.get(suf)
    if fam:
        o = _code_outline(text, fam)
        if o:
            return n, [f"{pad}{ln}" for ln in o.splitlines()]
    return n, []


# dir_outline 的固定排除（gitignore 之外的硬排除；不排 .agent——agent 定义目录本身常要列）
_DIR_OUTLINE_HARD = frozenset({".git", "__pycache__", "node_modules", ".venv", "venv",
                               ".mypy_cache", ".pytest_cache", ".ruff_cache", "dist", "build",
                               ".idea", ".vscode", ".next", ".cache"})


def dir_outline(path: str, max_files: int = 200, max_depth: int = 6) -> str:
    """目录大纲：树状列出目录内文件与各文件大纲（assembly DSL 的 dir: 装配项，也可直接调用）。
    .md → path + 各级标题行号树；代码文件（py/js/ts/cs/java/go/rs...）→ 类/函数定义行号；
    其余文件只列行数。忽略：workspace .gitignore 全部模式 + .git/__pycache__/node_modules 等
    固定清单 + 嵌套 git 仓库整棵剪枝。上限保护：max_files 文件数 / max_depth 深度 /
    单文件 512KB · 40 定义，超出截断并标注。"""
    target = _resolve(path)
    if not target.exists():
        return f"[路径不存在] {path}"
    if target.is_file():
        n, outline = _file_outline(target, "    ")
        return f"{path} [{n} 行]\n" + "\n".join(outline)
    # gitignore 过滤谓词（延迟 import：agent 顶部 import real_tools，防循环）
    from agent import _make_gitignore_filter
    keep_dir, keep_file = _make_gitignore_filter(WORKSPACE)
    out, nfiles, cut = [], 0, False

    def walk(d: Path, depth: int):
        nonlocal nfiles, cut
        if cut or depth > max_depth:
            return
        try:
            entries = sorted(d.iterdir(), key=lambda p: (p.is_file(), p.name.lower()))
        except OSError:
            return
        for p in entries:
            if cut:
                return
            rel = p.relative_to(WORKSPACE).as_posix()
            if p.is_dir():
                if p.name in _DIR_OUTLINE_HARD or p.name.endswith(".egg-info"):
                    continue
                if (p / ".git").exists():
                    continue               # 嵌套 git 仓库（coze-studio 等）整棵剪枝
                if not keep_dir(rel):
                    continue
                out.append(f"{'  ' * depth}{p.name}/")
                walk(p, depth + 1)
            else:
                if not keep_file(rel):
                    continue
                if nfiles >= max_files:
                    cut = True
                    return
                nfiles += 1
                n, outline = _file_outline(p, "  " * (depth + 2))
                out.append(f"{'  ' * depth}{p.name} [{n} 行]")
                out.extend(outline)
    walk(target, 0)
    if cut:
        out.append(f"… 文件数超 {max_files} 截断（收窄 path 或分目录列）")
    return "\n".join(out) or "(空目录)"


def concat_files(pattern: str, max_files: int = 50, max_chars: int = 64000) -> str:
    """按 glob 模式拼接多个文件内容（assembly DSL 的 tool: 项读多文件用，如 concat_files('.agent/rules/*.md')）。
    pattern: workspace 内相对 glob（* 单层 / ** 任意层；也接受目录名=目录下所有文件）。
    匹配结果按路径排序，每段以「=== <相对路径> ===」分隔；超 max_files 截断、单文件超 512KB 跳过。
    返回拼接文本；无匹配返回空串（调用方按需忽略）。"""
    import glob as _glob
    pat = (pattern or "").strip().strip('"').strip("'")
    if not pat:
        return ""
    base = WORKSPACE
    # 目录名：展开为目录下所有文件（不含隐藏/常见排除目录）
    if "*" not in pat and "?" not in pat and "[" not in pat:
        d = _resolve(pat) if not pat.startswith(("..", "/", "\\")) else (base / pat)
        if d.is_dir():
            pat = (str(d.relative_to(base)).replace("\\", "/") or ".") + "/**/*"
    matches = []
    for m in _glob.glob(str(base / pat), recursive=True):
        p = Path(m)
        if not p.is_file():
            continue
        if any(seg in {".git", "__pycache__", "node_modules", ".venv", "venv",
                       "dist", "build", ".mypy_cache"} for seg in p.parts):
            continue
        matches.append(p)
    matches.sort()
    if not matches:
        return ""
    out = []
    n = 0
    for p in matches[:max_files]:
        if p.stat().st_size > 512_000:
            continue
        try:
            txt = p.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        rel = p.relative_to(base).as_posix()
        out.append(f"=== {rel} ===\n{txt.rstrip()}")
        n += 1
    body = "\n\n".join(out)
    if len(matches) > max_files:
        body += f"\n\n… 共 {len(matches)} 个文件，超 max_files={max_files} 截断（收窄 pattern）"
    return body[:max_chars]


def web_search(query: str) -> str:
    """用 DuckDuckGo 搜索，返回前几条结果的标题/链接/摘要。无需 API key。
    注意：搜索引擎可能临时限流；国内网络下可能需要代理。
    返回 JSON 字符串 {success, count, result, error}——success 为结构化输出字段，
    工作流 plugin 节点可直接引用 web_search_node.success 判断成功与否。"""
    import json as _json
    import warnings
    out = {"success": False, "count": 0, "result": "", "error": ""}
    try:
        try:
            from ddgs import DDGS               # 新包名
        except ImportError:
            from duckduckgo_search import DDGS  # 旧包名（会触发重命名警告）
    except ImportError:
        out["error"] = "未安装搜索库（pip install ddgs）"
        return _json.dumps(out, ensure_ascii=False)
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")      # 屏蔽旧包的重命名警告
            results = list(DDGS().text(query, max_results=5))
    except Exception as e:
        out["error"] = f"{type(e).__name__}: {e}（搜索引擎可能限流，或国内需代理）"
        return _json.dumps(out, ensure_ascii=False)
    out["success"] = True
    out["count"] = len(results)
    if results:
        lines = [f"{i}. {r.get('title','')}\n   {(r.get('href') or r.get('link') or '')}\n   {(r.get('body') or r.get('snippet') or '')}"
                 for i, r in enumerate(results, 1)]
        out["result"] = "\n\n".join(lines)
    else:
        out["result"] = f"没有搜到关于 '{query}' 的结果"
    return _json.dumps(out, ensure_ascii=False)


def _html_to_text(html: str) -> tuple[str, str]:
    """HTML → (title, 正文文本)。剥 script/style，块级标签换行，压缩空白。标准库实现，零依赖。"""
    import re
    from html.parser import HTMLParser

    class _Extractor(HTMLParser):
        SKIP = {"script", "style", "noscript", "template", "svg", "iframe"}
        BLOCK = {"p", "div", "br", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6",
                 "section", "article", "header", "footer", "ul", "ol", "table",
                 "blockquote", "pre", "hr", "form", "nav", "aside"}

        def __init__(self):
            super().__init__(convert_charrefs=True)
            self.parts: list[str] = []
            self.title_parts: list[str] = []
            self._skip_depth = 0
            self._in_title = False

        def handle_starttag(self, tag, attrs):
            if tag in self.SKIP:
                self._skip_depth += 1
            elif tag == "title":
                self._in_title = True
            elif tag in self.BLOCK:
                self.parts.append("\n")

        def handle_endtag(self, tag):
            if tag in self.SKIP:
                self._skip_depth = max(0, self._skip_depth - 1)
            elif tag == "title":
                self._in_title = False
            elif tag in self.BLOCK:
                self.parts.append("\n")

        def handle_data(self, data):
            if self._skip_depth:
                return
            if self._in_title:
                self.title_parts.append(data)
            elif data.strip():
                self.parts.append(data)

    p = _Extractor()
    try:
        p.feed(html)
        p.close()
    except Exception:
        pass  # 残缺 HTML 也尽量用已解析的部分
    lines = [re.sub(r"[ \t　]+", " ", ln).strip() for ln in "".join(p.parts).splitlines()]
    return "".join(p.title_parts).strip(), "\n".join(ln for ln in lines if ln)


def open_url(url: str, start: int = 0, max_chars: int = 8000) -> str:
    """抓取网页并提取正文文本（HTML 剥标签；JSON/纯文本原样），支持分页续读。
    url: 网页地址（http/https）；start: 从第几个字符开始读（0-based，默认 0）；
    max_chars: 本次最多返回字符数（默认 8000）。返回头部含总字数，未读完时按提示传 start 续读。"""
    import requests
    if not url.lower().startswith(("http://", "https://")):
        url = "https://" + url
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                             "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"}
    try:
        resp = requests.get(url, headers=headers, timeout=15)
    except Exception as e:
        return f"[抓取失败] {type(e).__name__}: {e}\n（网络不通或国内需代理）"
    # header 未声明 charset 时 requests 默认 ISO-8859-1，中文页会乱码 → 用探测编码
    if not resp.encoding or resp.encoding.lower() == "iso-8859-1":
        resp.encoding = resp.apparent_encoding
    ctype = (resp.headers.get("Content-Type") or "").lower()
    body = resp.text or ""
    if "html" in ctype or (not ctype and body.lstrip()[:1] == "<"):
        title, text = _html_to_text(body)
    else:
        title, text = "", body  # JSON / 纯文本等直接原样
    total = len(text)
    if total == 0:
        return f"[{url} HTTP {resp.status_code}] （提取不到正文，可能是纯 JS 渲染页面）"
    start = max(0, int(start))
    if start >= total:
        return f"[越界] 正文共 {total} 字符，start={start} 超出范围"
    end = min(start + max(1, int(max_chars)), total)
    status = "" if resp.status_code == 200 else f" HTTP {resp.status_code} |"
    title_part = f" 标题:{title} |" if title else ""
    more = f"，续读传 start={end}" if end < total else "，已读完"
    return f"[{url}{status}{title_part} 第 {start}-{end-1} 字 / 共 {total} 字{more}]\n" + text[start:end]


def _paginate_text(text: str, label: str, start: int, max_chars: int) -> str:
    """对纯文本分页：返回带头部（第 X-Y 字 / 共 N 字）的切片。越界报错。"""
    total = len(text)
    start = max(0, int(start))
    if start >= total:
        return f"[越界] 共 {total} 字符，start={start} 超出范围"
    end = min(start + max(1, int(max_chars)), total)
    more = f"，续读传 start={end}" if end < total else "，已读完"
    return f"[{label} | 第 {start}-{end-1} 字 / 共 {total} 字{more}]\n" + text[start:end]


_WORKFLOW_SPEC_URL = "https://raw.githubusercontent.com/vgp7758/Agt/main/docs/workflow-spec.md"
_WORKFLOW_SPEC_LOCAL = Path(__file__).resolve().parent.parent / "docs" / "workflow-spec.md"


def read_workflow_spec(start: int = 0, max_chars: int = 6000) -> str:
    """读取工作流规范全文（docs/workflow-spec.md）。【写工作流前务必先读】了解节点类型/字段/变量引用。
    返回格式：<structure> 标题大纲（层级 + 行号）+ <content> 正文。
    从线上 git raw 读取（与本地 docs/ 同源），网络不通时回退本地 docs/。
    start: 在此标题目录中从第 start 个字符开始读正文(0-based)；
    max_chars: 本次最多返回字符数(默认 6000)。"""
    text = None
    # 1) 优先线上 git raw（保证拿到最新版）
    try:
        import requests
        r = requests.get(_WORKFLOW_SPEC_URL, headers={"User-Agent": "agt-agent"}, timeout=15)
        if r.status_code == 200 and r.text:
            text = r.text
    except Exception:
        pass
    # 2) 兜底本地 docs/（pip 安装后随包附带；开发期在仓库根）
    if text is None and _WORKFLOW_SPEC_LOCAL.exists():
        try:
            text = _WORKFLOW_SPEC_LOCAL.read_text(encoding="utf-8")
        except Exception:
            pass
    if not text:
        return f"[读取失败] git raw 与本地 {_WORKFLOW_SPEC_LOCAL} 均无法获取 workflow-spec.md"
    # .md 文件用 _md_snapshot：先输出 <structure> 标题大纲 + 行号，再 <content> 正文
    return _md_snapshot(text)


# 工作流 demo 读取（git raw，本地兜底）
_DEMO_BASE_URL = "https://raw.githubusercontent.com/vgp7758/Agt/main/.agent/workflows/"
_DEMO_LOCAL_DIR = Path(__file__).resolve().parent.parent / ".agent" / "workflows"
_DEMOS = {
    "composite_demo": "循环+批处理+单节点批处理三合一（迭代入口/continue/break 模型，多工具组合+本地变量+筛选 nth）",
    "full_demo": "全节点类型演示（意图分流→各分支处理→聚合→序列化，覆盖 15 种节点）",
}


def _fetch_demo_text(name: str) -> str:
    """从 git raw 读 demo XML，失败回退本地。"""
    text = None
    try:
        import requests
        r = requests.get(_DEMO_BASE_URL + name + ".xml", headers={"User-Agent": "agt-agent"}, timeout=15)
        if r.status_code == 200 and r.text:
            text = r.text
    except Exception:
        pass
    if text is None:
        local = _DEMO_LOCAL_DIR / (name + ".xml")
        if local.exists():
            try:
                text = local.read_text(encoding="utf-8")
            except Exception:
                pass
    return text


def _builtin_tools_reference() -> str:
    """列出工作流可用的内置工具（LIGHT_TOOLS + 外置脚本工具）及示例 plugin 节点 XML。"""
    lines = ["=== 工作流内置工具（未注册给 Agent，只能在工作流 plugin 节点用）===",
             "这些轻量工具（add/split/sleep 等）Agent 不能直接调用，仅工作流编排可用。",
             "调用：<node type=\"plugin\" toolName=\"工具名\">，输出 raw（工具返回值）。",
             "入参 <in> 接上游输出 ref=\"节点ID.字段\"，或字面量 literal=\"值\"。",
             ""]
    _TS = {int: "integer", float: "number", bool: "boolean", list: "list", dict: "object"}
    _all = list(LIGHT_TOOLS)
    try:   # 外置脚本工具一并列入（纯函数型已迁 tools/builtin/）
        from script_tools import scan_script_tools
        _all += list(scan_script_tools())
    except Exception:
        pass
    for t in _all:
        lines.append(f"【{t.name}】{t.description}")
        ins = []
        for pname, param in t._sig.parameters.items():
            ptype = t._hints.get(pname, str)
            ts = _TS.get(ptype, "string")
            ins.append(f'    <in name="{pname}" type="{ts}"/>')
        lines.append(f'  示例：<node id="N" type="plugin" toolName="{t.name}">')
        lines.extend(ins)
        lines.append(f'    <out name="raw" type="string"/>')
        lines.append('  </node>')
        lines.append("")
    return "\n".join(lines)


def read_workflow_demo(demo: str = "", start: int = 0, max_chars: int = 8000) -> str:
    """读取工作流 demo XML 示例，或列出可用 demo 清单。
    【写工作流前参考】了解循环/批处理/各节点的 XML 写法。
    demo: 空则返回 demo 清单；'composite_demo' 读循环+批处理三合一；'full_demo' 读全节点类型演示。
    start/max_chars: 读取指定 demo 时的分页（XML 较长可续读）。
    提示：先用 list_workflow_nodes 了解可用节点，再用 query_workflow_node 查节点 XML 示例。"""
    if demo:
        if demo not in _DEMOS:
            return f"[未知 demo] {demo}，可选：{', '.join(_DEMOS)}"
        text = _fetch_demo_text(demo)
        if not text:
            return f"[读取失败] {demo}.xml（git raw 与本地均不可用）"
        header = f"=== {demo}.xml —— {_DEMOS[demo]} ===\n"
        return header + _paginate_text(text, demo + ".xml", start, max_chars)
    # 无 demo：返回清单
    parts = ["=== 工作流 demo 清单（传 demo=名称 读取完整 XML）==="]
    for name, desc in _DEMOS.items():
        parts.append(f"  - {name}: {desc}")
    parts.append("")
    parts.append("提示：用 list_workflow_nodes 查看所有可用节点类型，用 query_workflow_node 查具体节点 XML 示例。")
    return "\n".join(parts)


# ===== 工作流节点目录 ====

# ===== 核心节点目录（插件节点的目录条目在各自 .py 的 _CATALOG 声明，动态聚合）=====
_CORE_NODE_CATALOG = [
    {
        "type": "1", "name": "开始 (Start)",
        "desc": "工作流入口，定义外部调用时需传入的参数（即工作流工具的函数签名）",
        "xml": "<!-- 开始节点：定义工作流入参。每个工作流有且仅有一个开始节点(id=100001) -->\n<node id=\"100001\" type=\"start\">\n  <out name=\"query\" type=\"string\" required=\"true\"/>\n  <out name=\"max_results\" type=\"integer\" required=\"false\">10</out>\n</node>\n<!--\n  输入schema（外部->工作流）：\n  | 字段        | 类型    | 必填 | 说明         |\n  | query       | string  | ✓    | 查询关键词   |\n  | max_results | integer |      | 最大结果数，默认 10 |\n-->",
    },
    {
        "type": "2", "name": "结束 (End)",
        "desc": "工作流出口，收集上游节点输出作为工作流返回值。支持两种模式：returnVariables（取指定字段）和 useAnswerContent（渲染模板文本）",
        "xml": "<!-- 结束节点(id=900001)：定义工作流返回值 -->\n<!-- 模式1：returnVariables —— 取上游节点输出字段，组装成结构化返回值 -->\n<node id=\"900001\" type=\"end\">\n  <out name=\"answer\" ref=\"130001.output\"/>\n  <out name=\"confidence\" ref=\"140001.score\"/>\n</node>\n\n<!-- 模式2：useAnswerContent —— 渲染一段模板文本作为单一返回值 -->\n<node id=\"900001\" type=\"end\" useAnswerContent=\"true\">\n  <content><![CDATA[回答：{{answer}}（置信度：{{confidence}}）]]></content>\n</node>\n<!--\n  输出schema（工作流->外部）：\n  returnVariables 模式下输出各 out 字段；useAnswerContent 模式下输出 {\"output\": \"渲染文本\"}\n-->",
    },
    {
        "type": "4", "name": "插件 (Plugin)",
        "desc": "调用内置轻量工具（add/split/sleep 等）或用户自定义 Python 工具，输入输出通过 in/out 声明",
        "xml": "<!-- 插件节点：调用内置工具或用户工具 -->\n<!-- 示例1：调用内置加法工具 -->\n<node id=\"140001\" type=\"plugin\" toolName=\"add\">\n  <in name=\"a\" type=\"number\" ref=\"130001.x\"/>\n  <in name=\"b\" type=\"number\" literal=\"5\"/>\n  <out name=\"raw\" type=\"string\"/>\n</node>\n\n<!-- 示例2：调用内置分割工具 -->\n<node id=\"140002\" type=\"plugin\" toolName=\"split\">\n  <in name=\"text\" type=\"string\" ref=\"130001.output\"/>\n  <in name=\"separator\" type=\"string\" literal=\",\"/>\n  <out name=\"raw\" type=\"string\"/>\n</node>\n\n<!-- 示例3：调用用户自定义工具（.agent/workflows/tools/xxx.py） -->\n<node id=\"140003\" type=\"plugin\" toolName=\"my_custom_tool\">\n  <in name=\"param1\" type=\"string\" ref=\"130001.output\"/>\n  <in name=\"param2\" type=\"integer\" literal=\"42\"/>\n  <out name=\"raw\" type=\"string\"/>\n</node>\n<!--\n  内置工具列表见末尾 _builtin_tools_reference() 输出。\n  自定义工具放 .agent/workflows/tools/*.py，顶层函数自动注册。\n  输出固定为 raw（工具返回值字符串）。\n-->",
    },
    {
        "type": "21", "name": "循环 (Loop)",
        "desc": "复合节点(blocks)，对数组迭代或按次数循环。体内可用 Break(19)/Continue(29)/LoopSetVariable(20)。三种模式：array(遍历数组)、count(固定次数)、infinite(无限循环)",
        "xml": "<!-- 循环节点(composite)：迭代执行体内 blocks -->\n<node id=\"200001\" type=\"loop\">\n  <!-- loopType: array(遍历数组) | count(固定次数) | infinite(无限) -->\n  <param name=\"loopType\" literal=\"array\">array</param>\n  <!-- loopCount: count 模式下的循环次数 -->\n  <param name=\"loopCount\" type=\"integer\">10</param>\n\n  <!-- array 模式：声明要遍历的数组 -->\n  <in name=\"items\" ref=\"170001.filtered_outputs\"/>\n\n  <!-- 循环变量（可选）：初始值，体内 LoopSetVariable 节点可读写 -->\n  <param name=\"accumulator\" type=\"integer\" initialValue=\"0\"/>\n\n  <!-- 体内子节点 blocks（inline canvas） -->\n  <blocks>\n    <!-- 体内可用的特殊节点：LoopSetVariable(20) 读写循环变量 -->\n    <node id=\"200010\" type=\"setvar\">\n      <left>accumulator</left>                             <!-- 循环变量名 -->\n      <right ref=\"200011.output\"/>                         <!-- 新值 -->\n    </node>\n\n    <!-- 体内 LLM 节点：通过 loop-item / loop-index 引用当前迭代元素和索引 -->\n    <node id=\"200011\" type=\"llm\">\n      <in name=\"item\" loop-item=\"true\"/>                   <!-- 当前迭代元素 -->\n      <in name=\"index\" loop-index=\"true\"/>                 <!-- 当前索引(0-based) -->\n      <param name=\"prompt\"><![CDATA[处理第 {{index}} 项：{{item}}]]></param>\n      <out name=\"output\" type=\"string\"/>\n    </node>\n\n    <!-- 条件退出：选择器判断后走 Break 端口 -->\n    <node id=\"200012\" type=\"selector\">\n      <in name=\"output\" ref=\"200011.output\"/>\n      <branch>\n        <condition operator=\"Contain\" logic=\"2\">\n          <left ref=\"200011.output\"/>\n          <right literal=\"STOP\"/>\n        </condition>\n        <!-- true 端口 → Break -->\n      </branch>\n    </node>\n\n    <!-- Break(19): 强制退出循环 -->\n    <node id=\"200013\" type=\"break\"/>\n    <!-- Continue(29): 跳过本次迭代，进入下一次 -->\n    <node id=\"200014\" type=\"continue\"/>\n\n    <node id=\"200015\" type=\"llm\">\n      <in name=\"item\" loop-item=\"true\"/>\n      <param name=\"prompt\"><![CDATA[正常处理：{{item}}]]></param>\n      <out name=\"output\" type=\"string\"/>\n    </node>\n  </blocks>\n\n  <out name=\"all_outputs\" type=\"list\"/>\n</node>\n<!--\n  体内子节点引用迭代元素：<in name=\"x\" loop-item=\"true\"/>，取当前 item\n  体内子节点引用迭代索引：<in name=\"i\" loop-index=\"true\"/>，取当前 index\n  Break(19): 放在 Selector 的 true/false 出口后，满足条件时退出循环\n  Continue(29): 放在 Selector 出口后，满足条件时跳过本次\n  LoopSetVariable(20): left=变量名, right=新值（可 ref 上游），读写循环累加变量\n  输出：all_outputs（每轮迭代的末端输出 list）、final_变量名（循环变量最终值）\n-->",
    },
    {
        "type": "20", "name": "循环变量 (LoopSetVariable)",
        "desc": "在循环体内读写循环累加变量（仅 Loop/Batch 体内有效），left=变量名，right=新值",
        "xml": "<!-- 循环变量设置节点(type=20)：仅 Loop 或 Batch 体内使用 -->\n<node id=\"200010\" type=\"setvar\">\n  <left>counter</left>              <!-- 变量名（在循环节点的 variableParameters 中声明） -->\n  <right ref=\"200009.output\"/>      <!-- 新值：ref 引用体内节点输出，或 literal 写死 -->\n</node>\n<!--\n  left: 变量名字符串（不是 ref）\n  right: 新值，ref=体内节点ID.字段 或 literal=\"值\"\n  变量的最终值会出现在循环节点的输出中（final_counter 等）\n-->",
    },
    {
        "type": "19", "name": "循环中断 (Break)",
        "desc": "在循环体内强制退出整个循环（仅 Loop/Batch 体内有效），通常放在 Selector 的某个条件出口后",
        "xml": "<!-- Break 节点(type=19)：仅 Loop 或 Batch 体内使用，无条件退出循环 -->\n<node id=\"200013\" type=\"break\"/>\n<!--\n  通常用法：Selector 判断某条件→true 端口→连到 Break\n  注意：Break 和 Continue 没有 in/out，只需声明节点本身\n-->",
    },
    {
        "type": "29", "name": "循环继续 (Continue)",
        "desc": "在循环体内跳过当前迭代进入下一轮（仅 Loop/Batch 体内有效），通常放在 Selector 出口后",
        "xml": "<!-- Continue 节点(type=29)：仅 Loop 或 Batch 体内使用，跳过本轮迭代 -->\n<node id=\"200014\" type=\"continue\"/>\n<!--\n  通常用法：Selector 判断某条件→true 端口→连到 Continue\n  注意：Break 和 Continue 没有 in/out，只需声明节点本身\n-->",
    },
    {
        "type": "28", "name": "批处理 (Batch)",
        "desc": "复合节点(blocks)，对数组逐元素并发执行体内逻辑，支持 batchSize/concurrentSize 控制并发度，输出聚合结果列表",
        "xml": "<!-- 批处理节点(composite)：逐元素并发执行体内 blocks -->\n<node id=\"210001\" type=\"batch\">\n  <!-- batchSize: 每批处理条数；concurrentSize: 并发数 -->\n  <param name=\"batchSize\" type=\"integer\">5</param>\n  <param name=\"concurrentSize\" type=\"integer\">3</param>\n\n  <!-- 输入：要批处理的数组 -->\n  <in name=\"items\" ref=\"170001.filtered_outputs\"/>\n\n  <blocks>\n    <!-- 体内节点：通过 loop-item / loop-index 引用当前元素和索引 -->\n    <node id=\"210010\" type=\"llm\">\n      <in name=\"item\" loop-item=\"true\"/>\n      <in name=\"index\" loop-index=\"true\"/>\n      <param name=\"prompt\"><![CDATA[处理第 {{index}} 项：{{item}}]]></param>\n      <out name=\"output\" type=\"string\"/>\n    </node>\n  </blocks>\n\n  <out name=\"all_outputs\" type=\"list\"/>\n  <out name=\"filtered_outputs\" type=\"list\"/>\n</node>\n<!--\n  体内引用：loop-item=\"true\" 取当前元素，loop-index=\"true\" 取当前索引\n  输出：all_outputs(所有结果list), filtered_outputs(过滤null后的结果), nth_output(第n个结果)\n  体内也支持 Break(19) 和 Continue(29)\n-->",
    },
    {
        "type": "9", "name": "子工作流 (SubWorkflow)",
        "desc": "调用另一个已注册的工作流作为子流程，传入参数、获取结构化返回值",
        "xml": "<!-- 子工作流节点：调用另一个工作流 -->\n<node id=\"260001\" type=\"subworkflow\">\n  <!-- workflow: 目标工作流名（.agent/workflows/ 下的文件名，不含扩展名） -->\n  <param name=\"workflow\" literal=\"greet\">greet</param>\n\n  <!-- 输入：传给子工作流的参数（对应子工作流开始节点的 out 声明） -->\n  <in name=\"name\" ref=\"130001.output\"/>\n\n  <!-- 输出：子工作流结束节点返回的字段 -->\n  <out name=\"greeting\" type=\"string\"/>\n  <out name=\"output\" type=\"string\"/>\n</node>\n<!--\n  workflow 参数：目标工作流的文件名（不含扩展名）\n  输入字段对应子工作流开始节点(100001)声明的 out\n  输出字段对应子工作流结束节点(900001)的 out\n-->",
    },
    {
        "type": "13", "name": "输出发送 (OutputEmitter)",
        "desc": "交互式输出：在工作流执行中途向外部发送消息（工具模式下仅收集输出，不会真正交互）",
        "xml": "<!-- 输出发送节点：向外部发送中间结果（工具模式下仅记录） -->\n<node id=\"270001\" type=\"output\">\n  <in name=\"message\" ref=\"130001.output\"/>\n  <in name=\"data\" ref=\"150001.result\"/>\n</node>\n<!--\n  交互模式下向用户发送消息；工具模式下输出被收集到 ctx.emitMessages\n  通常和 InputReceiver(30) 配对使用，实现\"中间输出-等待输入-继续执行\"\n-->",
    },
    {
        "type": "31", "name": "注释 (Comment)",
        "desc": "纯注释节点，不参与执行，用于在画布上添加说明文字",
        "xml": "<!-- 注释节点：不参与执行，仅用于画布标注 -->\n<node id=\"290001\" type=\"comment\">\n  <content>这里是对后续逻辑的说明，不会被执行</content>\n</node>\n<!-- 注释节点在扫描和执行时均被跳过，不会产生任何输出 -->",
    },
]

def _node_catalog() -> list:
    """全量节点目录 = 核心（本表）+ 插件（node_plugins.catalog_entries 动态聚合——
    插件的 desc/xml 示例跟实现走，用户级 .agent/nodes/ 的目录声明同样生效）。
    延迟 import：node_plugins.default_dirs 反向依赖本模块 WORKSPACE，顶层 import 会循环。"""
    try:
        from node_plugins import catalog_entries
        return _CORE_NODE_CATALOG + catalog_entries()
    except Exception:
        return list(_CORE_NODE_CATALOG)



def list_workflow_nodes() -> str:
    """列出工作流所有可用节点类型（名称 + 类型码 + 简介）。
    先调用它了解节点全景，再用 query_workflow_node(type="3") 或 query_workflow_node(name="LLM") 查具体某个节点的完整 XML 示例。"""
    cat = _node_catalog()
    lines = ["=== 工作流可用节点（共 {} 种）===".format(len(cat)),
             "{:<6} {:<20} {}".format("type", "名称", "简介"),
             "-" * 70]
    for n in cat:
        lines.append("{:<6} {:<20} {}".format(n["type"], n["name"], n["desc"]))
    lines.append("")
    lines.append('用法：query_workflow_node(type="3") 或 query_workflow_node(name="LLM") 查看某节点的完整 XML 示例。')
    return "\n".join(lines)


def query_workflow_node(type: str = "", name: str = "") -> str:
    """查询某个工作流节点的完整 XML 示例（含输入/输出 schema 和字段说明）。
    type: 节点类型码（如 "3" 表示 LLM）；name: 节点名称模糊匹配（如 "LLM" 或 "循环"）。
    二选一，type 优先。先用 list_workflow_nodes 查看所有可用节点类型。

    特殊：type="4" 或 name="plugin" 时额外列出所有可用内置工具及其参数。"""
    if type:
        matches = [n for n in _node_catalog() if n["type"] == type]
    elif name:
        nl = name.lower()
        matches = [n for n in _node_catalog() if nl in n["name"].lower() or nl in n["desc"].lower()]
    else:
        return '请提供 type 或 name 参数。先用 list_workflow_nodes 查看所有可用节点。\n示例：query_workflow_node(type="3") 或 query_workflow_node(name="循环")'

    if not matches:
        hint = "可用 type：" + ", ".join(sorted({n["type"] for n in _node_catalog()}, key=lambda x: int(x) if x.isdigit() else 9999))
        return f"[未匹配] type={type}, name={name}\n{hint}\n先用 list_workflow_nodes 查看所有可用节点。"

    parts = []
    for n in matches:
        parts.append(f"=== {n['name']}（type={n['type']}）===")
        parts.append(f"用途：{n['desc']}")
        parts.append("")
        parts.append("--- XML 示例 ---")
        parts.append(n["xml"])
        parts.append("")

        # plugin 节点额外列出内置工具
        if n["type"] == "4":
            parts.append("--- 可用内置工具（LIGHT_TOOLS）---")
            parts.append(_builtin_tools_reference())

    return "\n".join(parts)




def run_shell(command: str) -> str:
    """执行一条系统 shell 命令，实时流式输出。超时由 TOOL_TIMEOUT 控制（可用 set_tool_timeout 调大）。"""
    return _run_subprocess_streaming(command, "run_shell", shell=True)


def run_script(script: str, payload: str = "") -> str:
    """运行本地 Python 脚本并返回其 stdout——用于在工作流中执行自己写的处理脚本。
    script: 脚本路径（相对 workspace，如 'tools/analyze.py' 或 '.agent/workflows/tools/x.py'）；
    payload: 传给脚本的 JSON 负载，脚本通过环境变量 PAYLOAD 读取（json.loads 后使用）。
    【工作流用法】前置一个 ToJSON 节点把若干输入组装成 JSON，output 接本节点 payload；
    脚本约定：读 os.environ['PAYLOAD'] 取参数、print 输出结果（后续可接 FromJSON 解析）。"""
    import subprocess
    import sys
    import os
    target = _resolve(script)
    if not target.exists():
        return f"[脚本不存在] {script}（相对 workspace，如 tools/xxx.py）"
    if target.suffix.lower() not in (".py", ".pyw"):
        return f"[仅支持 .py 脚本] {script}"
    env = dict(os.environ)
    env["PAYLOAD"] = payload or ""
    # 把 workspace 加入 PYTHONPATH，让脚本能 import workspace 内其它模块（如 tools/ 下的辅助模块）
    pp = str(WORKSPACE)
    if env.get("PYTHONPATH"):
        pp = pp + os.pathsep + env["PYTHONPATH"]
    env["PYTHONPATH"] = pp
    try:
        proc = subprocess.run([sys.executable, str(target)], capture_output=True, text=True,
                              timeout=TOOL_TIMEOUT, env=env, cwd=str(WORKSPACE),
                              encoding="utf-8", errors="replace",
                              creationflags=(subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0))
    except subprocess.TimeoutExpired:
        return f"[脚本执行超时（>{TOOL_TIMEOUT}s），可用 set_tool_timeout 调大]"
    except Exception as e:
        return f"[执行失败] {type(e).__name__}: {e}"
    out = (proc.stdout or "").strip()
    if proc.returncode != 0:
        err = (proc.stderr or "").strip()
        return f"[脚本出错 rc={proc.returncode}]\nstderr: {err[-500:]}\nstdout: {out[-500:]}"
    return out or "(无输出)"


def set_tool_timeout(seconds: int) -> str:
    """设置 run_python / run_shell 的超时秒数（默认 10）。
    某些工具调用可能很长（如模拟、训练），可调大到 600（10分钟）甚至 1800（30分钟）。
    seconds: 超时秒数（1~7200）。"""
    global TOOL_TIMEOUT
    if not (1 <= seconds <= 7200):
        return f"❌ seconds 需在 1~7200 之间，收到 {seconds}"
    old = TOOL_TIMEOUT
    TOOL_TIMEOUT = seconds
    return f"✅ 工具超时已从 {old}s 改为 {seconds}s"


def get_tool_timeout() -> str:
    """查看当前 run_python / run_shell 的超时秒数。"""
    return f"当前工具超时：{TOOL_TIMEOUT}s"


# length/to_uppercase/to_lowercase → tools/builtin/str_tools.py（纯函数外置）
# ===== 工作流 ReAct 原语三件套 =====
# _WF_CTX 由 workflow._handle_plugin 在调用这三个工具时注入（llm=执行上下文的 LLMClient、
# tools=工具箱）。它们只在工作流 plugin 节点里有意义——Agent/代码节点直接调用会拿到错误提示。
_WF_CTX: dict = {"llm": None, "tools": None}
# 工作流内临时模型切换的 client 缓存：model 参数指定的 provider 名 → 独立 LLMClient。
# 与 utility_client 同款惰性缓存模式（一次创建进程内复用）；call_recorder 继承主 llm（llm_calls.jsonl 可观测）
_WF_ALT_LLM_CACHE: dict = {}


def _wf_llm_for(model: str):
    """按 model 参数取 client：空名 → 上下文 llm（None=无上下文，llm_call 自行报错）；
    指定名字 → 独立 client 缓存（无则建）——上下文 llm 为 None 也不妨碍按名建（此前
    `llm is None` 与 `not name` 合并判断，导致无上下文+指定名也返回 None，llm_call 直接报
    'no LLM context'——独立测试/agent=None 场景下指定 model 完全可用）。"""
    llm = _WF_CTX.get("llm")
    name = (model or "").strip()
    if not name:
        return llm
    if llm is not None and name == getattr(llm, "model_name", ""):
        return llm   # 与上下文同名 → 复用
    cli = _WF_ALT_LLM_CACHE.get(name)
    if cli is None:
        try:
            import config
            if name not in config.MODELS:
                return llm   # 未配置的 provider 名 → 回退上下文 llm（不炸工作流）
            from llm_client import LLMClient
            # model_name 构造（与 utility_client 同款，该路径已验证）；
            # 注意 LLMClient 对未知名构造不抛异常（惰性校验）——须先查 config.MODELS
            cli = LLMClient(model_name=name, max_retries=2)
            cli.call_recorder = getattr(llm, "call_recorder", None)   # 流水继承（/stats 可观测）
            _WF_ALT_LLM_CACHE[name] = cli
        except Exception:
            return llm   # 构造异常 → 回退上下文 llm（不炸工作流）
    return cli


def llm_call(messages: list, tools: list = None, temperature: float = None,
             enable_thinking: bool = None, model: str = "") -> str:
    """原生 LLM 调用（工作流 ReAct 原语）：完整 messages + tools schema → 完整回包。
    messages: OpenAI 格式消息数组（system/user/assistant/tool 均可）；
    tools: get_tool_schemas 输出的 function schema 数组（留空=纯文本对话）；
    model: 模型选择——models.json 的 provider 名（如 glm-official-1），留空=跟随 utility/主模型；
           独立 client 缓存复用（同 provider 多节点零开销），流水记 llm_calls.jsonl（scene=wf:llm_call）。
    返回 JSON（plugin 节点自动解析成对象，下游可 .content / .tool_calls.0.name 引用）：
      {content, reasoning, tool_calls:[{id,name,arguments}], usage, finish_reason, error}"""
    import json as _json
    llm = _wf_llm_for(model)
    if llm is None:
        return _json.dumps({"error": "llm_call 需在工作流 plugin 节点中执行（无 LLM 上下文）",
                            "content": "", "reasoning": "", "tool_calls": []}, ensure_ascii=False)
    overrides = {}
    if temperature is not None:
        overrides["temperature"] = float(temperature)
    if enable_thinking is not None:
        # XML literal 可能以字符串 "False"/"true" 传来（JSON 小写才解析成 bool）——按布尔语义归一
        _et = enable_thinking
        if isinstance(_et, str):
            _et = _et.strip().lower() in ("1", "true", "yes", "on")
        overrides["enable_thinking"] = bool(_et)
    try:
        resp = llm.chat(messages or [], tools=tools or None, scene="wf:llm_call", **overrides)
        return _json.dumps({
            "content": resp.content or "",
            "reasoning": resp.reasoning or "",
            "tool_calls": resp.tool_calls or [],
            "usage": resp.usage,
            "finish_reason": resp.finish_reason,
            "model": getattr(resp, "model", ""),
        }, ensure_ascii=False, default=str)
    except Exception as e:
        return _json.dumps({"error": f"{type(e).__name__}: {e}", "content": "",
                            "reasoning": "", "tool_calls": []}, ensure_ascii=False)


def get_tool_schemas(names: str = "") -> str:
    """获取工具 schema 数组（OpenAI function 格式），直接喂给 llm_call 的 tools 入参。
    names: 逗号分隔的工具名过滤（如 'read_file,grep'）；留空=全部工具。
    工作流 ReAct 循环的标准装配：get_tool_schemas → llm_call(tools=...) 。"""
    import json as _json
    tb = _WF_CTX.get("tools")
    if tb is None:
        return _json.dumps({"error": "get_tool_schemas 需在工作流 plugin 节点中执行"}, ensure_ascii=False)
    want = {n.strip() for n in (names or "").split(",") if n.strip()}
    # include_hidden=True：LIGHT_TOOLS 整箱 hidden（算术工具/三件套对主 LLM 不投影），
    # 但工作流 ReAct 需要它们的 schema——get_tool_schemas 的语义就是"全量工具表"
    schemas = [s for s in tb.schemas(include_hidden=True)
               if not want or s.get("function", {}).get("name") in want]
    return _json.dumps(schemas, ensure_ascii=False)


def call_tool(name: str, arguments: dict = None) -> str:
    """按名字动态执行工具箱中的工具（工作流 ReAct 原语：执行 llm_call 返回的 tool_calls）。
    name: 工具名（如 llm_call.tool_calls.0.name）；arguments: 参数 dict（如 ...tool_calls.0.arguments）。
    返回工具结果文本（错误也转文本，不炸工作流）。⚠️ 不递归拦截：工具集含 wf_* 时模型可能调到
    工作流自身造成递归——可用 get_tool_schemas(names=...) 过滤掉。"""
    tb = _WF_CTX.get("tools")
    if tb is None:
        return "[call_tool 需在工作流 plugin 节点中执行]"
    try:
        return tb.call((name or "").strip(), arguments or {})
    except Exception as e:
        return f"[工具 {name} 执行失败] {type(e).__name__}: {e}"


LLM_CALL_OUTPUTS = [
    {"name": "content", "type": "string", "description": "最终回答正文（无工具调用时）"},
    {"name": "reasoning", "type": "string", "description": "思考过程（推理模型）"},
    {"name": "tool_calls", "type": "list", "description": "工具调用数组；空=模型给出最终回答",
     "schema": [{"name": "id", "type": "string"}, {"name": "name", "type": "string"},
                {"name": "arguments", "type": "object"}]},
    {"name": "finish_reason", "type": "string", "description": "stop / tool_calls / length"},
    {"name": "usage", "type": "object", "description": "token 用量"},
    {"name": "error", "type": "string", "description": "错误信息（成功为空）"},
]



def git_commit(message: str, files: str = "") -> str:
    """git add + commit + push 一体（Agent 的标准提交通道）。
    自动在 commit message 末尾附加 trailer：Co-authored-by: Agt <vgp123@foxmail.com>
    （GitHub 识别为共同作者并在提交页展示——标记 AI 参与开发）。
    message: 提交信息（首行摘要 + 可选正文）；files: 要 add 的文件（逗号分隔，留空=git add -A 全部变更）。
    工作区无变更时跳过 commit/push 并提示；push 为联网操作，调用即视为本次授权。"""
    import subprocess
    if not (WORKSPACE / ".git").exists():
        return "[非 git 仓库] 当前 workspace 没有 .git"
    msg = ((message or "").strip() or "update") + "\n\nCo-authored-by: Agt <vgp123@foxmail.com>"

    def _git(*args, timeout=180):
        return subprocess.run(["git", *args], cwd=str(WORKSPACE), capture_output=True,
                              text=True, encoding="utf-8", errors="replace", timeout=timeout,
                              creationflags=(subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0))

    targets = [f.strip() for f in (files or "").split(",") if f.strip()]
    r_add = _git("add", *(targets if targets else ["-A"]))
    if r_add.returncode != 0:
        return f"[git add 失败] {(r_add.stderr or '').strip()[:300]}"
    r_ci = _git("commit", "-m", msg)
    out = (r_ci.stdout or "") + (r_ci.stderr or "")
    if r_ci.returncode != 0:
        if "nothing to commit" in out or "no changes added" in out:
            return "（无变更可提交——工作区干净，已跳过 commit/push）"
        return f"[git commit 失败] {out.strip()[:300]}"
    r_push = _git("push")
    if r_push.returncode != 0:
        return (f"✅ commit 成功但 push 失败：{(r_push.stderr or r_push.stdout).strip()[:300]}\n"
                f"（网络/权限问题可稍后手动 git push；commit 已在本地）")
    r_log = _git("log", "-1", "--oneline")
    return f"✅ 已提交并推送\n{r_log.stdout.strip()}\n（trailer: Co-authored-by: Agt）"


def _myers_diff(a_lines, b_lines):
    """Myers Diff 算法（纯算法函数，工具入口见 diff_files）。
    输入: 两个字符串列表（按行）。
    输出: 差异列表 [(action, line)]，action: '-' 删除(A侧行) / '+' 插入(B侧行) / ' ' 相同。"""
    A, B = a_lines, b_lines
    N, M = len(A), len(B)
    MAX = N + M
    V = {1: 0}
    trace = []
    for D in range(MAX + 1):
        trace.append(dict(V))          # trace[d] = d 轮开始前（=d-1 轮结束后）的 V 快照
        for k in range(-D, D + 1, 2):
            if k == -D or (k != D and V.get(k - 1, 0) < V.get(k + 1, 0)):
                x = V.get(k + 1, 0)      # 向下（插入）
            else:
                x = V.get(k - 1, 0) + 1  # 向右（删除）
            y = x - k
            while x < N and y < M and A[x] == B[y]:
                x += 1
                y += 1
            V[k] = x
            if x >= N and y >= M:
                return _myers_backtrack(trace, A, B, D)
    return []


def _myers_backtrack(trace, A, B, D):
    """回溯生成 diff。⚠️ prev_k 判断与 prev_x 取值必须用【同一层】快照 trace[d]（=d-1 轮结束状态，
    即第 d 步编辑的出发点）——取 trace[d-1] 会错一层，回溯偏离合法编辑链（重放错乱/B[y]越界）。"""
    x, y = len(A), len(B)
    result = []
    for d in range(D, 0, -1):
        V = trace[d]
        k = x - y
        if k == -d or (k != d and V.get(k - 1, 0) < V.get(k + 1, 0)):
            prev_k = k + 1
        else:
            prev_k = k - 1
        prev_x = V.get(prev_k, 0)      # ← 同一层（trace[d]）取，与 prev_k 判断一致
        prev_y = prev_x - prev_k
        while x > prev_x and y > prev_y:   # 对角线（相同段）从后往前收集
            x -= 1
            y -= 1
            result.append((' ', A[x]))
        if x > prev_x:                    # 水平步 = 删除 A 侧
            x -= 1
            result.append(('-', A[x]))
        elif y > prev_y:                  # 垂直步 = 插入 B 侧
            y -= 1
            result.append(('+', B[y]))
    while x > 0 and y > 0:                # d=0 纯对角线（起点前导相同段）
        x -= 1
        y -= 1
        result.append((' ', A[x]))
    result.reverse()
    return result


def _resolve_read(path: str) -> Path:
    """只读场景的路径解析：先按 workspace 沙箱解析；越界（绝对路径或 ../ 逃逸）时放行为
    直接路径——diff_files 等纯只读工具用（对比 workspace 外的备份/参照文件）。
    写操作仍走 _resolve 严格沙箱（读写不对称：读放行、写拦截）。"""
    try:
        return _resolve(path)
    except PermissionError:
        p = Path(path)
        return p if p.is_absolute() else (WORKSPACE / path).resolve()


def _render_unified_diff(A, B, ops, label1, label2, context, a_offset=0, b_offset=0):
    """公共渲染：Myers ops → unified diff 文本（@@ hunk + 带行号 -/+ 行）。diff_files/diff_lines 共用。
    a_offset/b_offset：分段对比时行号还原为文件内绝对行号（range_a[0]-1）——输出行号直接可用。"""
    dels = sum(1 for a, _ in ops if a == '-')
    adds = sum(1 for a, _ in ops if a == '+')
    if not dels and not adds:
        return f"[无差异] {label1} 与 {label2} 内容相同（{len(A)} 行）"
    annot, i1, i2 = [], 0, 0
    for act, ln in ops:
        if act in (' ', '-'):
            i1 += 1
        if act in (' ', '+'):
            i2 += 1
        annot.append((act, ln,
                      (i1 + a_offset) if act in (' ', '-') else None,
                      (i2 + b_offset) if act in (' ', '+') else None))
    ctx = max(0, int(context))
    changed_idx = [i for i, (a, *_r) in enumerate(annot) if a != ' ']
    hunks, s = [], 0
    for j in range(1, len(changed_idx) + 1):
        if j == len(changed_idx) or changed_idx[j] - changed_idx[j - 1] > 2 * ctx + 1:
            lo = max(0, changed_idx[s] - ctx)
            hi = min(len(annot) - 1, changed_idx[j - 1] + ctx)
            hunks.append((lo, hi))
            s = j
    parts = [f"[diff {label1} ({len(A)}行) vs {label2} ({len(B)}行) | -{dels} +{adds} | {len(hunks)} 处差异]"]
    for lo, hi in hunks:
        seg = annot[lo:hi + 1]
        a_start = next((a2 for _a, _l, a2, _b in seg if a2 is not None), 0)
        b_start = next((b2 for _a, _l, _a2, b2 in seg if b2 is not None), 0)
        a_n = sum(1 for _a, _l, a2, _b in seg if a2 is not None)
        b_n = sum(1 for _a, _l, _a2, b2 in seg if b2 is not None)
        parts.append(f"@@ -{a_start},{a_n} +{b_start},{b_n} @@")
        for act, ln, a2, b2 in seg:
            if act == ' ':
                parts.append(f"  {ln}")
            elif act == '-':
                parts.append(f"-{a2}│ {ln}")
            else:
                parts.append(f"+{b2}│ {ln}")
    out = "\n".join(parts)
    if len(out) > 20000:
        out = out[:20000] + f"\n...（输出截断，全量差异 -{dels} +{adds} 行；可减小 context 或分段对比）"
    return out


def _parse_range(rng, total, label):
    """[起,止] → (start_idx0, end_idx1, err)；成功时 err=None（只报错误，不夹带数据——
    此前成功返回 (a,b) 元组被调用方当 truthy 错误误判，整个 diff 结果变成一个 tuple）。"""
    if rng is None:
        return 0, total, None
    if not (isinstance(rng, list) and rng and all(isinstance(x, int) and not isinstance(x, bool) for x in rng)):
        return None, None, f"[参数错误] {label} 需为 [起,止] 整数数组，收到 {rng!r}"
    a = max(1, rng[0])
    b = min(total, rng[1]) if len(rng) > 1 and rng[1] else total
    if a > b:
        return None, None, f"[参数错误] {label}={rng} 越界或起>止（该文件共 {total} 行）"
    return a - 1, b, None


def diff_files(file1: str, file2: str, context: int = 2,
               range_a: list = None, range_b: list = None) -> str:
    """Myers Diff 对比两个文件，返回 unified diff 风格的差异（@@ hunk 头 + -/+ 行）。
    file1/file2: 路径（相对 workspace 或绝对路径；只读放行——可对比 workspace 外的备份/参照文件）。
    context: 每个 hunk 前后的上下文行数（默认 2）。
    range_a/range_b: 分段对比——各自文件的行范围 [起, 止]（1-based 含两端；如 [100, 200]）。
      两个文件行号错位时各传各的（file1 取 100-200 行 vs file2 取 120-220 行也行）；
      只传 range_a 时 range_b 默认同 range_a。输出行号仍是【文件内绝对行号】（可直接喂 edit）。
    头部含两侧行数/增删统计；完全相同返回「无差异」。适合对比 改前/改后、备份/当前。"""
    t1, t2 = _resolve_read(file1), _resolve_read(file2)
    if not t1.exists():
        return f"[文件不存在] {file1}"
    if not t2.exists():
        return f"[文件不存在] {file2}"
    A_all = t1.read_text(encoding="utf-8", errors="ignore").splitlines()
    B_all = t2.read_text(encoding="utf-8", errors="ignore").splitlines()
    if range_b is None and range_a is not None:
        range_b = list(range_a) if isinstance(range_a, list) else range_a
    s1, e1, err1 = _parse_range(range_a, len(A_all), "range_a")
    if err1:
        return err1
    s2, e2, err2 = _parse_range(range_b, len(B_all), "range_b")
    if err2:
        return err2
    A, B = A_all[s1:e1], B_all[s2:e2]
    seg1 = f" L{s1+1}-L{e1}/{len(A_all)}" if s1 or e1 < len(A_all) else ""
    seg2 = f" L{s2+1}-L{e2}/{len(B_all)}" if s2 or e2 < len(B_all) else ""
    return _render_unified_diff(A, B, _myers_diff(A, B),
                                file1 + seg1, file2 + seg2, context,
                                a_offset=s1, b_offset=s2)




# web_search 的结构化输出（success 作为字段，供工作流 plugin 节点引用判断成功与否）
WEB_SEARCH_OUTPUTS = [
    {"name": "success", "type": "boolean", "description": "搜索是否成功"},
    {"name": "count", "type": "integer", "description": "结果条数"},
    {"name": "result", "type": "string", "description": "格式化的结果文本（标题/链接/摘要）"},
    {"name": "error", "type": "string", "description": "失败原因（成功时为空）"},
]

REAL_TOOLS = Toolbox(
    Tool(run_python, param_descriptions={
        "code": "内联 Python 代码（多行，写临时文件再跑）。和 file 二选一。",
        "file": "已存在的 .py 文件路径（跑已保存的脚本传这个，别再用 subprocess 包壳）。和 code 二选一。",
    }),
    Tool(read_file, param_descriptions={
        "line_numbers": "默认 True=每行前加行号(宽度按本段最大行号自适应对齐)，便于接下来用 insert/delete/move 按行号编辑；False=纯文本",
    }),
    Tool(write_file),
    Tool(edit),
    Tool(insert, param_descriptions={
        "entries": "插入点数组，每项 {line: 1-based行号, content: 文本(可多行)}；在该行之前插入；line<=0或超过总行数=追加末尾",
        "version": "read_file/grep 返回的 file_version；不匹配说明文件已改、需重读",
    }),
    Tool(delete, param_descriptions={
        "start_line": "要删除的起始行号（1-based，含）",
        "end_line": "要删除的结束行号（1-based，含）",
        "version": "read_file/grep 返回的 file_version；不匹配说明文件已改、需重读",
    }),
    Tool(move, param_descriptions={
        "start_line": "要搬移的起始行号（1-based，含）",
        "end_line": "要搬移的结束行号（1-based，含）",
        "dst_line": "搬到这里之前（按原始行号理解）",
        "version": "read_file/grep 返回的 file_version；不匹配说明文件已改、需重读",
    }),
    Tool(replace_lines, param_descriptions={
        "entries": "替换段数组，每项 {range:[起,止](1-based含两端), content:新文本(可多行)}；[n,n]替换单行；content=\"\"删该范围；多处传原始行号即可(内部降序应用)",
        "version": "read_file/grep 返回的 file_version；不匹配说明文件已改、需重读",
    }),
    Tool(list_dir),
    Tool(grep, param_descriptions={
        "pattern": "搜索模式，默认按正则（支持 a|b 多选一、. * 等元字符，与 ripgrep 一致）",
        "regex": "True=按正则(默认)；False=按字面匹配（把 pattern 当普通字符串，不解释元字符）",
        "context": "每条命中前后各显示几行（默认 0=只显示匹配行）",
    }),
    Tool(find_function, param_descriptions={
        "name": "函数/方法名（精确匹配，非正则）",
        "path": "文件路径；也可传目录则扫描其下同语言文件（跨文件找定义）",
        "lang": "语言提示(python/js/ts/cs/java/cpp/go...)，不传按扩展名识别",
        "context": "函数体前后额外显示几行（默认 0）",
    }),
    Tool(web_search, outputs=WEB_SEARCH_OUTPUTS),
    Tool(open_url),
    Tool(read_workflow_spec),
    Tool(read_workflow_demo),
    Tool(list_workflow_nodes),
    Tool(query_workflow_node),
    Tool(run_shell),
    Tool(run_script),
    Tool(set_tool_timeout),
    Tool(get_tool_timeout),
    Tool(diff_files, param_descriptions={
        "context": "每个 hunk 前后的上下文行数（默认 2）",
        "range_a": "file1 的行范围 [起,止]（1-based 含两端，如 [100,200]）——大文件分段对比用；不传=全文",
        "range_b": "file2 的行范围；不传时默认同 range_a。两文件行号错位时各传各的",
    }),
    Tool(git_commit, param_descriptions={
        "message": "提交信息（首行摘要+可选正文）；末尾自动附加 Co-authored-by: Agt trailer",
        "files": "要 add 的文件（逗号分隔）；留空=git add -A 全部变更",
    }),
)

LIGHT_TOOLS = Toolbox(
    # 纯函数型（length/to_uppercase/to_lowercase）→ tools/builtin/str_tools.py
    # kv_cache_read/write → tools/builtin/kv_tools.py（_KV_CACHE 状态随外置件走）
    # diff_lines → tools/builtin/diff_tools.py（Myers 三件套副本；框架侧 diff_files 用本文件算法）
    # cosine_sim/emb_probe → 本体 rag.py，注册 tools/builtin/rag_tools.py（与 embedder 单例共生）
    # 工作流 ReAct 原语三件套（仅工作流 plugin 节点可用；执行时由 workflow 引擎注入 llm/tools 上下文）
    Tool(llm_call, outputs=LLM_CALL_OUTPUTS, param_descriptions={
        "messages": "OpenAI 格式消息数组（system/user/assistant/tool 均可；循环中用循环变量累积）",
        "tools": "工具 schema 数组（接 get_tool_schemas 的输出）；留空=纯文本对话",
    }),
    Tool(get_tool_schemas, param_descriptions={
        "names": "逗号分隔的工具名过滤（如 'read_file,grep'）；留空=全部工具（⚠️ 含 wf_* 递归风险，ReAct 建议显式列出）",
    }),
    Tool(call_tool, param_descriptions={
        "name": "工具名（接 llm_call.tool_calls.0.name）",
        "arguments": "参数 dict（接 llm_call.tool_calls.0.arguments）",
    }),
    Tool(dir_outline, param_descriptions={
        "path": "要列大纲的目录（workspace 内；也可传单个文件）",
        "max_files": "最多展开的文件数（默认 200，超出截断标注）",
        "max_depth": "最大下钻深度（默认 6）",
    }),
    Tool(concat_files, param_descriptions={
        "pattern": "glob 模式（如 .agent/rules/*.md；目录名=该目录全部文件）",
        "max_files": "最多拼接的文件数（默认 50）",
        "max_chars": "拼接结果字符上限（默认 64000）",
    }),
    hidden=True,
)

# 全部内置工具（编辑器 /api/tools 返回这个）
ALL_BUILTIN_TOOLS = Toolbox(*(list(REAL_TOOLS) + list(LIGHT_TOOLS)))


def infer_tool_outputs(tool) -> list[dict]:
    """从工具的返回值类型注解推断输出 schema。
    str→[{name:'result',type:'string'}], float→number, int→integer, bool→boolean,
    list→list, dict→object, 无注解或无返回值→[{name:'raw',type:'string'}]。"""
    try:
        hints = getattr(tool.func, "__annotations__", {})
    except Exception:
        hints = {}
    ret = hints.get("return")
    mapping = {"str": "string", "int": "integer", "float": "number", "number": "number",
               "bool": "boolean", "list": "list", "dict": "object",
               str: "string", int: "integer", float: "number", bool: "boolean",
               list: "list", dict: "object"}
    if ret is None or ret is type(None):
        return [{"name": "raw", "type": "string", "description": "工具返回"}]
    # ret 可能是字符串（from __future__）或类型
    key = ret if isinstance(ret, str) else (ret.__name__ if hasattr(ret, "__name__") else str(ret))
    if ret in mapping:
        return [{"name": "result", "type": mapping[ret], "description": "工具返回值"}]
    if key in mapping:
        return [{"name": "result", "type": mapping[key], "description": "工具返回值"}]
    return [{"name": "raw", "type": "string", "description": "工具返回"}]


def make_autonomous_tools(agent) -> list:
    """生成绑定到指定 Agent 的纯自主模式工具。"""
    from datetime import datetime, timedelta

    def set_autonomous_mode(end_time: str = None, duration_minutes: int = None,
                            prompt: str = None, goal_check_code: str = None) -> str:
        """开启纯自主模式：任务完成后自动继续工作，直到时间到或目标达成（哪个先满足）。
        end_time: 结束时间 "HH:MM"（今天）或 "YYYY-MM-DD HH:MM"；
        duration_minutes: 持续分钟数（如 180=3小时）；
        goal_check_code: 目标验证 Python 脚本（print('PASS')=达成→自动停止）；
        prompt: 自动继续时的提示词。
        以上四个参数至少提供一个（end_time/duration_minutes/goal_check_code 三选一即可）。"""
        from datetime import datetime, timedelta
        try:
            # 目标脚本
            if goal_check_code:
                agent.goal_check_script = goal_check_code.strip()
            # 结束时间
            target = None
            if duration_minutes is not None:
                target = datetime.now() + timedelta(minutes=int(duration_minutes))
            elif end_time:
                end_time = end_time.strip()
                try:
                    target = datetime.strptime(end_time, "%Y-%m-%d %H:%M")
                except ValueError:
                    today = datetime.now().date()
                    target = datetime.strptime(f"{today} {end_time}", "%Y-%m-%d %H:%M")
                    if target < datetime.now():
                        target += timedelta(days=1)
            if target is None and not goal_check_code:
                return "[参数缺失] 至少提供 end_time / duration_minutes / goal_check_code 之一"
            agent.set_autonomous_mode(target or datetime.max, prompt)
            parts = []
            if target:
                parts.append(f"持续到 {target.strftime('%Y-%m-%d %H:%M')}")
            if goal_check_code:
                parts.append("目标验证脚本已设置")
            return f"✅ 纯自主模式已开启（{'，'.join(parts)}）"
        except Exception as e:
            return f"[开启失败] {type(e).__name__}: {e}"

    def exit_autonomous_mode() -> str:
        """退出纯自主模式。"""
        agent.exit_autonomous_mode()
        return "✅ 纯自主模式已关闭"

    def autonomous_status() -> str:
        """查看纯自主模式当前状态。"""
        if not agent.autonomous_mode:
            return "纯自主模式：未开启"
        if agent.is_autonomous_active():
            return f"纯自主模式：已开启，持续到 {agent.autonomous_end_time.strftime('%Y-%m-%d %H:%M')}\n" \
                   f"自动提示词：{agent.autonomous_prompt}\n" \
                   f"待处理消息队列：{len(agent.pending_messages)} 条"
        else:
            return "纯自主模式：已超时（自动关闭）"

    def set_goal_check(script: str) -> str:
        """设置目标达成验证脚本（Python）。自主循环每轮结束后跑它：输出 'PASS' 表示目标达成、自动结束自主模式；
        否则继续。如：拉坦克天梯分 ≥ 3000 → print('PASS')。"""
        if not script or not script.strip():
            return "[错误] script 不能为空"
        agent.goal_check_script = script.strip()
        return "✅ 目标验证脚本已设置（自主循环每轮结束后自动检查）"

    def check_goal() -> str:
        """手动运行一次目标验证脚本，返回输出（PASS=达成/FAIL=未达成/空=未设目标）。"""
        if not agent.goal_check_script:
            return "(未设置目标验证脚本，用 set_goal_check(script) 设置)"
        return agent.run_goal_check() or "(空输出)"

    return [Tool(set_autonomous_mode), Tool(exit_autonomous_mode), Tool(autonomous_status),
            Tool(set_goal_check), Tool(check_goal)]
